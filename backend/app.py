import os
import re

from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from typing import Optional

import auth
import db
import notifications
import pdfs_reparaciones
import pdfs_rh
import pdfs_equipos
import calendario_ics
try:
    import microsip
    MICROSIP_DISPONIBLE = True
except ImportError:
    # El driver 'fdb' todavía no está en requirements.txt / instalado en el
    # servidor — no tronamos toda la app por esto, solo dejamos claro el
    # motivo si alguien intenta usar las rutas de Microsip mientras tanto.
    microsip = None
    MICROSIP_DISPONIBLE = False

db.init_db()

app = FastAPI(title="Tickets TI — Multiempresa")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


# ---- Dependencias de autorización ----

def requiere_superadmin(usuario: dict = Depends(auth.get_current_user)) -> dict:
    if usuario["rol"] != "superadmin":
        raise HTTPException(status_code=403, detail="Solo el super administrador puede hacer esto")
    return usuario


def requiere_empresa_o_master(usuario: dict = Depends(auth.get_current_user)) -> dict:
    """Como requiere_empresa, pero sin excluir al rol 'master' — solo se usa en los
    dos endpoints a los que master sí tiene acceso: /api/meta y /api/dashboard."""
    if usuario["rol"] == "superadmin" or not usuario.get("empresa_id"):
        raise HTTPException(status_code=403, detail="Esta acción es solo para usuarios de una empresa")
    return usuario


def requiere_empresa(usuario: dict = Depends(requiere_empresa_o_master)) -> dict:
    """Cualquier rol de una empresa EXCEPTO 'master' (solo Dashboard). El rol
    'almacen' YA NO está bloqueado de fondo aquí — por defecto solo tiene
    Reparaciones (como siempre), pero el administrador puede darle acceso a
    otros módulos desde Administrar → Accesos, igual que a cualquier otro rol."""
    if usuario["rol"] == "master":
        raise HTTPException(status_code=403, detail="Tu usuario solo tiene acceso al Dashboard")
    return usuario


def requiere_empresa_o_almacen(usuario: dict = Depends(requiere_empresa_o_master)) -> dict:
    """Como requiere_empresa, pero también permite al rol 'almacen' — se usa solo en
    los endpoints de Reparaciones a los que un encargado de almacén sí tiene acceso
    (ver reparaciones y firmar la recepción en su propia sucursal)."""
    if usuario["rol"] == "master":
        raise HTTPException(status_code=403, detail="Tu usuario solo tiene acceso al Dashboard")
    return usuario


def requiere_dashboard(usuario: dict = Depends(requiere_empresa_o_master)) -> dict:
    if usuario["rol"] not in ("admin", "master"):
        raise HTTPException(status_code=403, detail="No tienes acceso al Dashboard")
    if usuario["rol"] == "admin":
        usuario = _con_permisos(usuario)
        if not usuario.get("acceso_dashboard", True):
            raise HTTPException(status_code=403, detail="No tienes acceso al Dashboard")
    return usuario


def requiere_admin(usuario: dict = Depends(requiere_empresa)) -> dict:
    if usuario["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo el administrador de tu empresa puede hacer esto")
    return usuario


def requiere_staff(usuario: dict = Depends(requiere_empresa)) -> dict:
    if usuario["rol"] not in ("admin", "tecnico"):
        raise HTTPException(status_code=403, detail="No tienes permiso para hacer esto")
    return usuario


def _con_permisos(usuario: dict) -> dict:
    """Agrega al dict del usuario sus permisos vigentes (leídos frescos de la base,
    no del JWT) — aplica a cualquier rol de empresa, para que el administrador pueda
    restringir módulos a técnicos y empleados, igual que ya podía hacerlo consigo
    mismo entre distintos administradores."""
    if usuario["rol"] in ("admin", "tecnico", "usuario", "almacen", "encargado_sucursal"):
        permisos = db.obtener_permisos_usuario(usuario["id"])
        usuario = {**usuario, **permisos}
    return usuario


def requiere_acceso_equipos(usuario: dict = Depends(requiere_empresa)) -> dict:
    """Antes exigía ser admin/técnico (requiere_staff). Ahora también deja pasar
    a 'almacen' si el administrador le dio el acceso — el rol 'usuario' (empleado)
    se sigue quedando fuera, igual que siempre."""
    if usuario["rol"] not in ("admin", "tecnico", "almacen"):
        raise HTTPException(status_code=403, detail="No tienes permiso para hacer esto")
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_equipos", True):
        raise HTTPException(status_code=403, detail="No tienes acceso al módulo de Equipos")
    return usuario


def requiere_acceso_compras(usuario: dict = Depends(requiere_staff)) -> dict:
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_compras", True):
        raise HTTPException(status_code=403, detail="No tienes acceso a administrar Compras")
    return usuario


def requiere_admin_compras(usuario: dict = Depends(requiere_admin)) -> dict:
    """Exclusivo del administrador: catálogo de artículos (crear/editar/dar de baja)
    y programar ciclos de compra. El técnico ya no gestiona ninguno de los dos,
    solo puede pedir del catálogo como cualquier empleado."""
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_compras", True):
        raise HTTPException(status_code=403, detail="No tienes acceso a administrar Compras")
    return usuario


def requiere_admin_rh(usuario: dict = Depends(requiere_admin)) -> dict:
    """Aprobar o rechazar una incidencia de Recursos Humanos es exclusivo del
    administrador — cualquier persona puede levantar su propia incidencia, pero
    solo el admin decide."""
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_rh", True):
        raise HTTPException(status_code=403, detail="No tienes acceso a administrar Recursos Humanos")
    return usuario


def requiere_ver_compras(usuario: dict = Depends(requiere_empresa)) -> dict:
    """Como requiere_acceso_compras, pero para las rutas que también usan técnicos y
    empleados (ver catálogo, ver ciclos, hacer un pedido) — ahora respeta la
    restricción configurada para CUALQUIER rol, no solo administrador."""
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_compras", True):
        raise HTTPException(status_code=403, detail="No tienes acceso al módulo de Compras")
    return usuario


def requiere_ver_rh(usuario: dict = Depends(requiere_empresa)) -> dict:
    """Igual que requiere_ver_compras, pero para Recursos Humanos: cualquier
    persona puede levantar su propia incidencia salvo que el administrador le
    haya quitado el acceso al módulo entero."""
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_rh", True):
        raise HTTPException(status_code=403, detail="No tienes acceso al módulo de Recursos Humanos")
    return usuario


def requiere_ver_tickets(usuario: dict = Depends(requiere_empresa)) -> dict:
    """Igual que requiere_ver_compras/requiere_ver_rh, pero para el tablero de
    Tickets — el módulo por defecto, que ahora también se le puede quitar a
    cualquier persona desde Administrar → Usuarios."""
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_tickets", True):
        raise HTTPException(status_code=403, detail="No tienes acceso al módulo de Tickets")
    return usuario


def requiere_ver_reparaciones(usuario: dict = Depends(requiere_empresa_o_almacen)) -> dict:
    """Igual que requiere_ver_tickets, pero para Reparaciones. El rol 'almacen'
    siempre tiene acceso — es su único módulo, no se le puede quitar — los
    demás roles sí pueden perder este acceso desde Administrar → Accesos."""
    if usuario["rol"] == "almacen":
        return usuario
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_reparaciones", True):
        raise HTTPException(status_code=403, detail="No tienes acceso al módulo de Reparaciones")
    return usuario


def requiere_admin_completo(usuario: dict = Depends(requiere_admin)) -> dict:
    usuario = _con_permisos(usuario)
    if not usuario.get("acceso_administracion", True):
        raise HTTPException(status_code=403, detail="No tienes acceso al módulo de Administración")
    return usuario


# ==================== AUTENTICACIÓN ====================

class LoginPayload(BaseModel):
    username: str
    password: str


@app.post("/api/auth/login")
def login(payload: LoginPayload):
    usuario = db.obtener_usuario_por_username(payload.username)
    if not usuario or not usuario["activo"] or not auth.verificar_password(payload.password, usuario["password_hash"]):
        raise HTTPException(status_code=401, detail="Usuario o contraseña incorrectos")
    token = auth.crear_token(usuario)
    return {
        "token": token,
        "usuario": {
            "id": usuario["id"], "username": usuario["username"],
            "nombre": usuario["nombre_completo"], "rol": usuario["rol"],
            "empresa_id": usuario["empresa_id"],
        },
    }


@app.get("/api/auth/me")
def me(usuario: dict = Depends(auth.get_current_user)):
    return usuario


class CambiarMiContrasena(BaseModel):
    password_actual: str
    password_nueva: str = Field(min_length=6)


@app.post("/api/auth/cambiar-password")
def cambiar_mi_password(payload: CambiarMiContrasena, usuario: dict = Depends(auth.get_current_user)):
    registro = db.obtener_usuario_por_username(usuario["username"])
    if not registro or not auth.verificar_password(payload.password_actual, registro["password_hash"]):
        raise HTTPException(status_code=401, detail="Tu contraseña actual no es correcta")
    db.actualizar_usuario(usuario["id"], password=payload.password_nueva)
    return {"ok": True}


# ==================== EMPRESAS (superadmin) ====================

class NuevaEmpresa(BaseModel):
    nombre: str = Field(min_length=1, max_length=120)
    admin_username: str = Field(min_length=3, max_length=40)
    admin_password: str = Field(min_length=6)
    admin_nombre: str = Field(min_length=1, max_length=120)


class ActualizacionEmpresa(BaseModel):
    nombre: Optional[str] = None
    activo: Optional[bool] = None


class NuevoLogo(BaseModel):
    logo_base64: str = Field(min_length=100)


@app.get("/api/empresas")
def listar_empresas(_: dict = Depends(requiere_superadmin)):
    return db.listar_empresas()


@app.post("/api/empresas")
def crear_empresa(payload: NuevaEmpresa, _: dict = Depends(requiere_superadmin)):
    if db.obtener_usuario_por_username(payload.admin_username):
        raise HTTPException(status_code=400, detail="Ese nombre de usuario ya está en uso")
    empresa_id = db.crear_empresa(payload.nombre.strip(), payload.admin_username.strip(), payload.admin_password, payload.admin_nombre.strip())
    return db.obtener_empresa(empresa_id)


@app.patch("/api/empresas/{empresa_id}")
def actualizar_empresa(empresa_id: int, payload: ActualizacionEmpresa, _: dict = Depends(requiere_superadmin)):
    if not db.obtener_empresa(empresa_id):
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    db.actualizar_empresa(empresa_id, payload.nombre, payload.activo)
    return db.obtener_empresa(empresa_id)


@app.post("/api/empresas/{empresa_id}/logo")
def subir_logo(empresa_id: int, payload: NuevoLogo, _: dict = Depends(requiere_superadmin)):
    if not db.obtener_empresa(empresa_id):
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    db.actualizar_logo_empresa(empresa_id, payload.logo_base64)
    return db.obtener_empresa(empresa_id)


class ClonarEmpresa(BaseModel):
    nombre_nueva_empresa: str = Field(min_length=1, max_length=120)
    sufijo_usuarios: str = Field(min_length=1, max_length=20, pattern=r"^[a-zA-Z0-9_-]+$")


@app.get("/api/empresas/{empresa_id}/respaldo")
def respaldo_empresa(empresa_id: int, _: dict = Depends(requiere_superadmin)):
    datos = db.exportar_empresa(empresa_id)
    if not datos:
        raise HTTPException(status_code=404, detail="Empresa no encontrada")

    import json
    from datetime import datetime as dt

    contenido = json.dumps(datos, indent=2, ensure_ascii=False, default=str)
    nombre_archivo = f"respaldo_{datos['empresa']['nombre'].replace(' ', '_')}_{dt.now().strftime('%Y%m%d')}.json"
    return Response(
        content=contenido, media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


@app.post("/api/empresas/{empresa_id}/clonar")
def clonar_empresa(empresa_id: int, payload: ClonarEmpresa, _: dict = Depends(requiere_superadmin)):
    if not db.obtener_empresa(empresa_id):
        raise HTTPException(status_code=404, detail="Empresa no encontrada")
    resultado = db.clonar_empresa(empresa_id, payload.nombre_nueva_empresa.strip(), payload.sufijo_usuarios.strip())
    if not resultado:
        raise HTTPException(status_code=400, detail="No se pudo clonar la empresa")
    return resultado


# ==================== POLÍTICAS DE LA EMPRESA (Reparaciones) ====================

DEFAULT_POLITICAS_TEXTO = """TÉRMINOS Y CONDICIONES DEL SERVICIO

1. GARANTÍA: El equipo cuenta con garantía únicamente si así se indica expresamente en la orden de servicio. Fuera de ese caso, NO EXISTE GARANTÍA sobre la reparación realizada ni sobre las refacciones utilizadas.

2. RESPALDO DE INFORMACIÓN: La empresa no se hace responsable por la pérdida de información, datos, programas o configuraciones almacenadas en el equipo. Es responsabilidad exclusiva del cliente respaldar su información antes de dejar el equipo a revisión.

3. TIEMPO DE RESGUARDO: Una vez que se notifica al cliente que el equipo está listo para su entrega, cuenta con 30 días naturales para recogerlo. Pasado ese plazo, la empresa no se hace responsable por el estado o resguardo del equipo.

4. DIAGNÓSTICO: El diagnóstico inicial puede variar una vez que el equipo es abierto. Cualquier cambio en el costo o alcance del servicio será notificado al cliente para su autorización antes de continuar con el trabajo.

5. ACCESORIOS: Solo se garantiza la devolución de los accesorios expresamente anotados en la orden de servicio al momento de la recepción.

Al firmar de conformidad, el cliente declara haber leído y estar de acuerdo con los términos aquí descritos."""


class ActualizacionPoliticas(BaseModel):
    texto: str = Field(min_length=1)


@app.get("/api/politicas")
def api_obtener_politicas(usuario: dict = Depends(requiere_empresa_o_almacen)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    texto = (empresa.get("politicas_texto") if empresa else None) or DEFAULT_POLITICAS_TEXTO
    return {"texto": texto}


@app.patch("/api/politicas")
def api_actualizar_politicas(payload: ActualizacionPoliticas, usuario: dict = Depends(requiere_admin)):
    db.actualizar_politicas_empresa(usuario["empresa_id"], payload.texto)
    return {"ok": True}


class ActualizacionApariencia(BaseModel):
    tema: Optional[str] = None
    color_acento: Optional[str] = None
    fondo_color: Optional[str] = None
    fondo_base64: Optional[str] = None


@app.patch("/api/apariencia")
def api_actualizar_apariencia(payload: ActualizacionApariencia, usuario: dict = Depends(requiere_admin)):
    enviados = payload.dict(exclude_unset=True)
    if payload.tema is not None and payload.tema not in ("oscuro", "claro"):
        raise HTTPException(status_code=400, detail="Tema inválido (usa 'oscuro' o 'claro')")
    patron_hex = re.compile(r"^#[0-9A-Fa-f]{6}$")
    if "color_acento" in enviados and payload.color_acento and not patron_hex.match(payload.color_acento):
        raise HTTPException(status_code=400, detail="El color de acento debe ser un código hexadecimal, ej. #D8192F")
    if "fondo_color" in enviados and payload.fondo_color and not patron_hex.match(payload.fondo_color):
        raise HTTPException(status_code=400, detail="El color de fondo debe ser un código hexadecimal, ej. #1A1B1D")
    if payload.fondo_base64 and len(payload.fondo_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La imagen de fondo pesa demasiado (máximo 5MB)")

    kwargs = {}
    if "color_acento" in enviados:
        kwargs["color_acento"] = payload.color_acento
    if "fondo_color" in enviados:
        kwargs["fondo_color"] = payload.fondo_color
    if "fondo_base64" in enviados:
        kwargs["fondo_base64"] = payload.fondo_base64
    db.actualizar_apariencia_empresa(usuario["empresa_id"], tema=payload.tema, **kwargs)
    return {"ok": True}


# ==================== META (para usuarios de una empresa) ====================

@app.get("/api/meta")
def meta(usuario: dict = Depends(requiere_empresa_o_master)):
    usuario = _con_permisos(usuario)
    empresa = db.obtener_empresa(usuario["empresa_id"])
    es_admin = usuario["rol"] == "admin"
    return {
        "estados": db.ESTADOS, "prioridades": db.PRIORIDADES, "roles": ["admin", "tecnico", "usuario", "master", "almacen", "encargado_sucursal"],
        "departamentos": [d["nombre"] for d in db.listar_departamentos(usuario["empresa_id"])],
        "categorias": [c["nombre"] for c in db.listar_categorias(usuario["empresa_id"])],
        "empresa_nombre": empresa["nombre"] if empresa else "",
        "empresa_logo": empresa["logo_base64"] if empresa else None,
        "apariencia": {
            "tema": (empresa.get("tema") if empresa else None) or "oscuro",
            "color_acento": empresa.get("color_acento") if empresa else None,
            "fondo_color": empresa.get("fondo_color") if empresa else None,
            "fondo_base64": empresa.get("fondo_base64") if empresa else None,
        },
        "tipos_equipo": db.TIPOS_EQUIPO, "estados_equipo": db.ESTADOS_EQUIPO,
        "tipos_mantenimiento": db.TIPOS_MANTENIMIENTO, "frecuencias_mantenimiento": db.FRECUENCIAS_MANTENIMIENTO,
        "estados_proyecto": db.ESTADOS_PROYECTO,
        "estados_tarea_proyecto": db.ESTADOS_TAREA_PROYECTO,
        "frecuencias_compra": db.FRECUENCIAS_COMPRA, "estados_ciclo_compra": db.ESTADOS_CICLO_COMPRA,
        "estados_reparacion": db.ESTADOS_REPARACION,
        "tipos_incidencia_rh": db.TIPOS_INCIDENCIA_RH, "estados_incidencia_rh": db.ESTADOS_INCIDENCIA_RH,
        "tipos_movimiento_horas_rh": db.TIPOS_MOVIMIENTO_HORAS_RH,
        "tablas_borrado_masivo": [{"key": k, "etiqueta": v["etiqueta"]} for k, v in db.TABLAS_BORRADO_MASIVO.items()],
        "mis_permisos": {
            "acceso_equipos": usuario.get("acceso_equipos", True),
            "acceso_administracion": usuario.get("acceso_administracion", True) if es_admin else False,
            "acceso_compras": usuario.get("acceso_compras", True),
            "acceso_rh": usuario.get("acceso_rh", True),
            "acceso_tickets": usuario.get("acceso_tickets", True),
            "acceso_reparaciones": True if usuario["rol"] == "almacen" else usuario.get("acceso_reparaciones", True),
            "acceso_dashboard": usuario.get("acceso_dashboard", True) if es_admin else True,
            "restriccion_categoria": usuario.get("restriccion_categoria") if es_admin else None,
        },
        "mi_departamento": db.obtener_departamento_usuario(usuario["id"]) if usuario["rol"] != "master" else None,
        "mi_sucursal_id": db.obtener_sucursal_id_usuario(usuario["id"]) if usuario["rol"] != "master" else None,
    }


# ==================== CALENDARIO PERSONAL (.ics) ====================
# Enlace de suscripción para el Calendario de iPhone o Google Calendar de
# Android — se suscriben UNA vez y de ahí en adelante los proyectos y
# mantenimientos que les toquen aparecen solos, sin volver a hacer nada.

@app.get("/api/mi-calendario")
def api_mi_calendario(usuario: dict = Depends(requiere_empresa)):
    token = db.obtener_o_crear_token_calendario(usuario["id"])
    return {"ruta": f"/calendario/{token}.ics"}


@app.post("/api/mi-calendario/regenerar")
def api_regenerar_mi_calendario(usuario: dict = Depends(requiere_empresa)):
    """Por si alguien comparte su enlace sin querer y quiere invalidarlo —
    esto rompe la suscripción vieja; hay que volver a suscribirse con la nueva."""
    token = db.regenerar_token_calendario(usuario["id"])
    return {"ruta": f"/calendario/{token}.ics"}


@app.get("/calendario/{token}.ics")
def calendario_ics_publico(token: str):
    """SIN autenticación normal a propósito — el Calendario de un celular no
    puede iniciar sesión ni mandar un token de sesión; el propio token (largo,
    aleatorio, imposible de adivinar) es la credencial, igual que hacen los
    calendarios de suscripción de Google Calendar, Notion, Trello, etc."""
    usuario = db.obtener_usuario_por_token_calendario(token)
    if not usuario:
        raise HTTPException(status_code=404, detail="Enlace de calendario no válido")
    empresa = db.obtener_empresa(usuario["empresa_id"])
    eventos = db.listar_eventos_calendario_usuario(usuario["empresa_id"], usuario["id"], usuario["rol"])
    contenido = calendario_ics.generar_ics(empresa["nombre"] if empresa else "Mark Inc", eventos)
    return Response(content=contenido, media_type="text/calendar; charset=utf-8",
                     headers={"Content-Disposition": "inline; filename=calendario.ics"})


@app.get("/api/notificaciones")
def api_obtener_notificaciones(usuario: dict = Depends(requiere_empresa_o_almacen)):
    """Resumen en vivo de pendientes (tickets, proyectos, ciclos de compra abiertos,
    etc.) — accesible a cualquier rol de la empresa, cada quien ve solo lo suyo.
    Cada notificación trae el id del elemento exacto, para poder ir directo a él."""
    if usuario["rol"] in ("master",):
        return []
    usuario = _con_permisos(usuario)
    acceso_compras = usuario.get("acceso_compras", True) if usuario["rol"] == "admin" else True
    acceso_rh = usuario.get("acceso_rh", True) if usuario["rol"] == "admin" else True
    acceso_reparaciones = True if usuario["rol"] == "almacen" else usuario.get("acceso_reparaciones", True)
    acceso_equipos = usuario.get("acceso_equipos", True)
    return db.obtener_notificaciones(usuario["empresa_id"], usuario["id"], usuario["rol"],
                                      acceso_compras, acceso_rh, acceso_reparaciones, acceso_equipos)


@app.get("/api/dashboard")
def api_dashboard(usuario: dict = Depends(requiere_dashboard)):
    return db.estadisticas_dashboard(usuario["empresa_id"])


NOMBRES_ESTADO_TICKET_PDF = {"abierto": "Abierto", "en_progreso": "En progreso", "resuelto": "Resuelto", "cerrado": "Cerrado"}
NOMBRES_ESTADO_REPARACION_BITACORA = {
    "nueva": "Reparación nueva en camino", "en_diagnostico": "Recibido en diagnóstico",
    "esperando_autorizacion": "Esperando autorización",
    "en_reparacion": "En reparación", "con_proveedor": "Con proveedor", "esperando_refaccion": "Esperando refacción",
    "control_calidad": "Control de calidad", "envio_sucursal": "Envío a sucursal", "en_traslado": "En traslado",
    "listo_entrega": "Listo para entrega", "entregado": "Entregado", "cancelado": "Cancelado",
}


@app.get("/api/dashboard/reporte.pdf")
def api_dashboard_reporte_pdf(usuario: dict = Depends(requiere_dashboard)):
    from io import BytesIO
    from datetime import datetime as dt

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable

    empresa = db.obtener_empresa(usuario["empresa_id"])
    d = db.detalle_dashboard(usuario["empresa_id"])

    ROJO = colors.HexColor("#D8192F")
    GRIS_CLARO = colors.HexColor("#F2F2F2")

    def tabla_resumen(por_estado, nombres):
        filas = [["Estado", "Cantidad"]] + [[nombres.get(e, e), str(n)] for e, n in por_estado.items()]
        t = Table(filas, colWidths=[9 * cm, 4 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), ROJO), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
        ]))
        return t

    def tabla_desglose(titulo_col, pares, styles):
        total = sum(n for _, n in pares) or 1
        filas = [[titulo_col, "Cantidad", "%"]] + [
            [nombre or "—", str(n), f"{n / total * 100:.0f}%"] for nombre, n in pares
        ]
        t = Table(filas, colWidths=[8 * cm, 3 * cm, 2 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#74767A")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey), ("ALIGN", (1, 0), (2, -1), "CENTER"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, GRIS_CLARO]),
        ]))
        return t

    styles = getSampleStyleSheet()
    estilo_seccion = ParagraphStyle("Seccion", parent=styles["Heading2"], fontSize=13, textColor=ROJO, spaceBefore=16, spaceAfter=6)
    estilo_sub = ParagraphStyle("Sub", parent=styles["Heading3"], fontSize=10.5, textColor=colors.HexColor("#333333"), spaceBefore=10, spaceAfter=4)

    elementos = [
        Paragraph(f"Dashboard General — {empresa['nombre'] if empresa else ''}", styles["Title"]),
        Paragraph(f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
        HRFlowable(width="100%", thickness=1, color=ROJO, spaceBefore=8, spaceAfter=10),
    ]

    # ---- Tickets ----
    elementos.append(Paragraph(f"Tickets — {d['tickets']['total']} en total", estilo_seccion))
    elementos.append(tabla_resumen(d["tickets"]["por_estado"], NOMBRES_ESTADO_TICKET_PDF))
    elementos.append(Paragraph("Desglose por departamento", estilo_sub))
    elementos.append(tabla_desglose("Departamento", d["tickets"]["por_departamento"], styles))
    elementos.append(Paragraph("Desglose por categoría", estilo_sub))
    elementos.append(tabla_desglose("Categoría", d["tickets"]["por_categoria"], styles))

    # ---- Reparaciones ----
    elementos.append(Paragraph(f"Reparaciones — {d['reparaciones']['total']} en total", estilo_seccion))
    elementos.append(tabla_resumen(d["reparaciones"]["por_estado"], NOMBRES_ESTADO_REPARACION_PDF))
    elementos.append(Paragraph("Desglose por sucursal", estilo_sub))
    elementos.append(tabla_desglose("Sucursal", d["reparaciones"]["por_sucursal"], styles))
    elementos.append(Paragraph("Desglose por cliente (doctor)", estilo_sub))
    elementos.append(tabla_desglose("Cliente", d["reparaciones"]["por_cliente"], styles))

    # ---- Proyectos ----
    elementos.append(Paragraph(f"Proyectos — {d['proyectos']['total']} en total", estilo_seccion))
    elementos.append(tabla_resumen(d["proyectos"]["por_estado"], NOMBRES_ESTADO_PROYECTO_PDF))

    # ---- Equipos ----
    elementos.append(Paragraph(f"Equipos — {d['equipos']['total']} en total", estilo_seccion))
    elementos.append(tabla_resumen(d["equipos"]["por_estado"], NOMBRES_ESTADO_EQUIPO))
    elementos.append(Paragraph("Desglose por tipo", estilo_sub))
    elementos.append(tabla_desglose("Tipo de equipo", [(NOMBRES_TIPO_EQUIPO.get(t, t), n) for t, n in d["equipos"]["por_tipo"]], styles))

    # ---- Compras ----
    elementos.append(Paragraph(f"Compras — {d['compras']['ciclos_total']} ciclo(s) en total", estilo_seccion))
    elementos.append(tabla_resumen(d["compras"]["ciclos_por_estado"], NOMBRES_ESTADO_CICLO_PDF))
    elementos.append(Paragraph(f"{d['compras']['pedidos_total']} pedido(s) registrados en total, en todos los ciclos.", styles["Normal"]))

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.8 * cm, rightMargin=1.8 * cm)
    doc.build(elementos)
    buffer.seek(0)
    nombre_archivo = f"dashboard_{dt.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.read(), media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


# ---- Reportes individuales por módulo, accesibles desde el Dashboard ----
# (mismo permiso que el Dashboard: admin o master — el usuario master no puede
# entrar a cada módulo por su cuenta, pero desde aquí sí puede bajar su PDF)

def _tabla_reporte_generica(encabezados, filas, color_header="#D8192F"):
    from reportlab.lib import colors
    from reportlab.platypus import Table, TableStyle
    datos = [encabezados] + filas
    t = Table(datos, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(color_header)),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    return t


def _armar_pdf_simple(titulo, empresa_nombre, elementos_extra, nombre_archivo):
    from io import BytesIO
    from datetime import datetime as dt
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph(f"{titulo} — {empresa_nombre}", styles["Title"]),
        Paragraph(f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]),
        Spacer(1, 16),
    ] + elementos_extra
    doc.build(elementos)
    buffer.seek(0)
    return Response(content=buffer.read(), media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}_{dt.now().strftime('%Y%m%d')}.pdf"})


@app.get("/api/dashboard/tickets.pdf")
def api_dashboard_tickets_pdf(usuario: dict = Depends(requiere_dashboard)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    tickets = db.listar_tickets(usuario["empresa_id"])
    filas = [[t["folio"], t["departamento"], t["solicitante_nombre"], t["prioridad"],
              NOMBRES_ESTADO_TICKET_PDF.get(t["estado"], t["estado"]), t["creado_en"][:16].replace("T", " ")] for t in tickets]
    tabla = _tabla_reporte_generica(["Folio", "Departamento", "Solicitante", "Prioridad", "Estado", "Creado"], filas)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph
    elementos = [Paragraph(f"{len(tickets)} ticket(s) en total", getSampleStyleSheet()["Heading2"]), tabla]
    return _armar_pdf_simple("Reporte de Tickets", empresa["nombre"] if empresa else "", elementos, "tickets")


@app.get("/api/dashboard/reparaciones.pdf")
def api_dashboard_reparaciones_pdf(usuario: dict = Depends(requiere_dashboard)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    reparaciones = db.listar_reparaciones(usuario["empresa_id"])
    filas = [[r["folio"], r.get("sucursal_nombre") or "—", r["cliente_nombre"], r.get("equipo") or "—",
              NOMBRES_ESTADO_REPARACION_PDF.get(r["estado"], r["estado"]), r.get("tecnico_nombre") or "sin asignar",
              f"${r.get('costo_total', 0):,.2f}"] for r in reparaciones]
    tabla = _tabla_reporte_generica(["Folio", "Sucursal", "Cliente", "Equipo", "Estado", "Técnico", "Costo total"], filas)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph
    elementos = [Paragraph(f"{len(reparaciones)} reparación(es) en total", getSampleStyleSheet()["Heading2"]), tabla]
    return _armar_pdf_simple("Reporte de Reparaciones", empresa["nombre"] if empresa else "", elementos, "reparaciones")


@app.get("/api/dashboard/proyectos.pdf")
def api_dashboard_proyectos_pdf(usuario: dict = Depends(requiere_dashboard)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    proyectos = db.listar_proyectos(usuario["empresa_id"], None)
    filas = []
    for p in proyectos:
        participantes = ", ".join(
            [u["nombre_completo"] for u in p.get("participantes_usuarios", [])] + p.get("participantes_departamentos", [])
        ) or "—"
        filas.append([p["nombre"], NOMBRES_ESTADO_PROYECTO_PDF.get(p["estado"], p["estado"]), participantes])
    tabla = _tabla_reporte_generica(["Nombre", "Estado", "Participantes"], filas)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph
    elementos = [Paragraph(f"{len(proyectos)} proyecto(s) en total", getSampleStyleSheet()["Heading2"]), tabla]
    return _armar_pdf_simple("Reporte de Proyectos", empresa["nombre"] if empresa else "", elementos, "proyectos")


@app.get("/api/dashboard/equipos.pdf")
def api_dashboard_equipos_pdf(usuario: dict = Depends(requiere_dashboard)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    equipos = db.listar_equipos(usuario["empresa_id"])
    filas = [[e["nombre"], NOMBRES_TIPO_EQUIPO.get(e["tipo"], e["tipo"]),
              " / ".join(filter(None, [e.get("marca"), e.get("modelo")])) or "—",
              e.get("responsable") or "—", NOMBRES_ESTADO_EQUIPO.get(e["estado"], e["estado"])] for e in equipos]
    tabla = _tabla_reporte_generica(["Nombre", "Tipo", "Marca/Modelo", "Responsable", "Estado"], filas)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph
    elementos = [Paragraph(f"{len(equipos)} equipo(s) en total", getSampleStyleSheet()["Heading2"]), tabla]
    return _armar_pdf_simple("Reporte de Equipos", empresa["nombre"] if empresa else "", elementos, "equipos")


@app.get("/api/dashboard/compras.pdf")
def api_dashboard_compras_pdf(usuario: dict = Depends(requiere_dashboard)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    ciclos = db.listar_ciclos_compra(usuario["empresa_id"])
    filas = [[c["nombre"], NOMBRES_FRECUENCIA_COMPRA_PDF.get(c["frecuencia"], c["frecuencia"]),
              NOMBRES_ESTADO_CICLO_PDF.get(c["estado"], c["estado"]), c["fecha_programada"][:10]] for c in ciclos]
    tabla = _tabla_reporte_generica(["Ciclo", "Frecuencia", "Estado", "Fecha programada"], filas)
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph
    elementos = [Paragraph(f"{len(ciclos)} ciclo(s) en total", getSampleStyleSheet()["Heading2"]), tabla]
    return _armar_pdf_simple("Reporte de Compras", empresa["nombre"] if empresa else "", elementos, "compras")


@app.get("/api/dashboard/rh.pdf")
def api_dashboard_rh_pdf(usuario: dict = Depends(requiere_dashboard)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    elementos = _elementos_reporte_rh_por_empleado(usuario["empresa_id"])
    return _armar_pdf_simple("Reporte de Recursos Humanos", empresa["nombre"] if empresa else "", elementos, "incidencias_rh")


# ==================== USUARIOS (dentro de la empresa, admin) ====================

class NuevoUsuario(BaseModel):
    username: str = Field(min_length=3, max_length=40)
    password: str = Field(min_length=6)
    nombre_completo: str = Field(min_length=1, max_length=120)
    rol: str
    telefono_whatsapp: Optional[str] = None
    puesto: Optional[str] = None
    sucursal_id: Optional[int] = None
    numero_empleado: Optional[str] = None


class ActualizacionUsuario(BaseModel):
    nombre_completo: Optional[str] = None
    rol: Optional[str] = None
    telefono_whatsapp: Optional[str] = None
    activo: Optional[bool] = None
    password: Optional[str] = Field(default=None, min_length=6)
    puesto: Optional[str] = None
    restriccion_categoria: Optional[str] = None
    acceso_equipos: Optional[bool] = None
    acceso_administracion: Optional[bool] = None
    acceso_compras: Optional[bool] = None
    acceso_rh: Optional[bool] = None
    acceso_dashboard: Optional[bool] = None
    acceso_tickets: Optional[bool] = None
    acceso_reparaciones: Optional[bool] = None
    sucursal_id: Optional[int] = None
    numero_empleado: Optional[str] = None


@app.get("/api/usuarios")
def api_listar_usuarios(usuario: dict = Depends(requiere_admin_completo)):
    return db.listar_usuarios(usuario["empresa_id"])


@app.get("/api/usuarios/tecnicos")
def api_listar_tecnicos(usuario: dict = Depends(requiere_empresa)):
    return db.listar_tecnicos_activos(usuario["empresa_id"])


@app.get("/api/usuarios/todos")
def api_listar_usuarios_activos(usuario: dict = Depends(requiere_empresa)):
    return db.listar_usuarios_activos(usuario["empresa_id"])


@app.post("/api/usuarios")
def api_crear_usuario(payload: NuevoUsuario, admin: dict = Depends(requiere_admin_completo)):
    if payload.rol not in ("admin", "tecnico", "usuario", "master", "almacen", "encargado_sucursal"):
        raise HTTPException(status_code=400, detail="Rol inválido")
    if db.obtener_usuario_por_username(payload.username):
        raise HTTPException(status_code=400, detail="Ese nombre de usuario ya está en uso")
    if payload.sucursal_id and not db.obtener_sucursal_reparacion(admin["empresa_id"], payload.sucursal_id):
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    uid = db.crear_usuario(admin["empresa_id"], payload.username, payload.password, payload.nombre_completo,
                            payload.rol, payload.telefono_whatsapp, payload.puesto, payload.sucursal_id,
                            payload.numero_empleado)
    return {"id": uid}


@app.patch("/api/usuarios/{usuario_id}")
def api_actualizar_usuario(usuario_id: int, payload: ActualizacionUsuario, admin: dict = Depends(requiere_admin_completo)):
    objetivo = next((u for u in db.listar_usuarios(admin["empresa_id"]) if u["id"] == usuario_id), None)
    if not objetivo:
        raise HTTPException(status_code=404, detail="Usuario no encontrado en tu empresa")
    if payload.rol and payload.rol not in ("admin", "tecnico", "usuario", "master", "almacen", "encargado_sucursal"):
        raise HTTPException(status_code=400, detail="Rol inválido")
    if payload.sucursal_id and not db.obtener_sucursal_reparacion(admin["empresa_id"], payload.sucursal_id):
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")

    enviados = payload.dict(exclude_unset=True)
    kwargs_extra = {}
    if "restriccion_categoria" in enviados:
        kwargs_extra["restriccion_categoria"] = payload.restriccion_categoria  # puede ser None para quitarla
    if "sucursal_id" in enviados:
        kwargs_extra["sucursal_id"] = payload.sucursal_id  # puede ser None para quitarla
    if "numero_empleado" in enviados:
        kwargs_extra["numero_empleado"] = payload.numero_empleado  # puede ser None para quitarlo

    db.actualizar_usuario(usuario_id, payload.nombre_completo, payload.rol, payload.telefono_whatsapp,
                           payload.activo, payload.password, payload.puesto,
                           acceso_equipos=payload.acceso_equipos, acceso_administracion=payload.acceso_administracion,
                           acceso_compras=payload.acceso_compras, acceso_rh=payload.acceso_rh,
                           acceso_dashboard=payload.acceso_dashboard, acceso_tickets=payload.acceso_tickets,
                           acceso_reparaciones=payload.acceso_reparaciones, **kwargs_extra)
    return {"ok": True}


@app.delete("/api/usuarios/{usuario_id}")
def api_eliminar_usuario(usuario_id: int, admin: dict = Depends(requiere_admin_completo)):
    if usuario_id == admin["id"]:
        raise HTTPException(status_code=400, detail="No puedes desactivarte a ti mismo")
    objetivo = next((u for u in db.listar_usuarios(admin["empresa_id"]) if u["id"] == usuario_id), None)
    if not objetivo:
        raise HTTPException(status_code=404, detail="Usuario no encontrado en tu empresa")
    db.eliminar_usuario(usuario_id)
    return {"ok": True}


# ==================== DEPARTAMENTOS Y CATEGORÍAS (admin) ====================

class NuevoNombre(BaseModel):
    nombre: str = Field(min_length=1, max_length=60)


class CambioEstado(BaseModel):
    activo: bool


@app.get("/api/departamentos")
def api_listar_departamentos(usuario: dict = Depends(requiere_admin_completo)):
    return db.listar_departamentos(usuario["empresa_id"], solo_activos=False)


@app.post("/api/departamentos")
def api_crear_departamento(payload: NuevoNombre, usuario: dict = Depends(requiere_admin_completo)):
    try:
        db.crear_departamento(usuario["empresa_id"], payload.nombre.strip())
    except Exception:
        raise HTTPException(status_code=400, detail="Ese departamento ya existe")
    return {"ok": True}


@app.patch("/api/departamentos/{depto_id}")
def api_cambiar_estado_departamento(depto_id: int, payload: CambioEstado, usuario: dict = Depends(requiere_admin_completo)):
    db.cambiar_estado_departamento(usuario["empresa_id"], depto_id, payload.activo)
    return {"ok": True}


@app.get("/api/categorias")
def api_listar_categorias(usuario: dict = Depends(requiere_admin_completo)):
    return db.listar_categorias(usuario["empresa_id"], solo_activos=False)


@app.post("/api/categorias")
def api_crear_categoria(payload: NuevoNombre, usuario: dict = Depends(requiere_admin_completo)):
    try:
        db.crear_categoria(usuario["empresa_id"], payload.nombre.strip().lower())
    except Exception:
        raise HTTPException(status_code=400, detail="Esa categoría ya existe")
    return {"ok": True}


@app.patch("/api/categorias/{cat_id}")
def api_cambiar_estado_categoria(cat_id: int, payload: CambioEstado, usuario: dict = Depends(requiere_admin_completo)):
    db.cambiar_estado_categoria(usuario["empresa_id"], cat_id, payload.activo)
    return {"ok": True}


class AsignarTecnicoCategoria(BaseModel):
    tecnico_id: Optional[int] = None  # None quita la auto-asignación


@app.patch("/api/categorias/{cat_id}/tecnico")
def api_asignar_tecnico_categoria(cat_id: int, payload: AsignarTecnicoCategoria, usuario: dict = Depends(requiere_admin_completo)):
    if payload.tecnico_id:
        tecnicos_validos = {t["id"] for t in db.listar_tecnicos_activos(usuario["empresa_id"])}
        if payload.tecnico_id not in tecnicos_validos:
            raise HTTPException(status_code=400, detail="Ese técnico no existe o no está activo en tu empresa")
    db.asignar_tecnico_categoria(usuario["empresa_id"], cat_id, payload.tecnico_id)
    return {"ok": True}


# ==================== TICKETS ====================

class NuevoTicket(BaseModel):
    departamento: str
    descripcion: str = Field(min_length=3)
    categoria: str
    prioridad: str = "media"


class ActualizacionTicket(BaseModel):
    estado: Optional[str] = None
    prioridad: Optional[str] = None
    asignado_a_id: Optional[int] = None


class NuevoComentario(BaseModel):
    texto: str = Field(min_length=1)
    archivo_base64: Optional[str] = None
    archivo_nombre: Optional[str] = None
    archivo_tipo: Optional[str] = None


class NuevaFirma(BaseModel):
    firma: str = Field(min_length=100)
    firmado_por: str = Field(min_length=1, max_length=120)


def _puede_ver_ticket(usuario, ticket):
    if usuario["rol"] == "usuario":
        return ticket["solicitante_id"] == usuario["id"]
    if usuario["rol"] == "tecnico":
        return ticket["asignado_a_id"] == usuario["id"]
    if usuario["rol"] == "admin":
        restriccion = usuario.get("restriccion_categoria")
        if restriccion:
            return ticket["categoria"] == restriccion
        return True
    return True


@app.get("/api/tickets")
def api_listar_tickets(estado: Optional[str] = None, prioridad: Optional[str] = None, categoria: Optional[str] = None,
                        departamento: Optional[str] = None, fecha_desde: Optional[str] = None, fecha_hasta: Optional[str] = None,
                        tecnico_id: Optional[int] = None, usuario: dict = Depends(requiere_ver_tickets)):
    solicitante_id = usuario["id"] if usuario["rol"] == "usuario" else None
    asignado_a_id = usuario["id"] if usuario["rol"] == "tecnico" else tecnico_id
    if usuario["rol"] == "admin" and usuario.get("restriccion_categoria"):
        categoria = usuario["restriccion_categoria"]  # se impone, ignora lo que haya mandado el cliente
    return db.listar_tickets(usuario["empresa_id"], estado, prioridad, categoria, solicitante_id,
                              departamento, fecha_desde, fecha_hasta, asignado_a_id)


# IMPORTANTE: esta ruta va ANTES de /api/tickets/{ticket_id} — FastAPI
# revisa las rutas en orden, y si {ticket_id} fuera primero,
# "reporte.pdf" se interpretaría como un intento de ticket_id (numérico)
# y fallaría con 422 antes de llegar aquí.
@app.get("/api/tickets/reporte.pdf")
def reporte_pdf(estado: Optional[str] = None, prioridad: Optional[str] = None, categoria: Optional[str] = None,
                 departamento: Optional[str] = None, fecha_desde: Optional[str] = None, fecha_hasta: Optional[str] = None,
                 tecnico_id: Optional[int] = None, usuario: dict = Depends(requiere_staff)):
    from io import BytesIO
    from datetime import datetime as dt

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    empresa = db.obtener_empresa(usuario["empresa_id"])
    usuario = _con_permisos(usuario)
    asignado_a_id = usuario["id"] if usuario["rol"] == "tecnico" else tecnico_id
    if usuario["rol"] == "admin" and usuario.get("restriccion_categoria"):
        categoria = usuario["restriccion_categoria"]
    todos = db.listar_tickets(usuario["empresa_id"], estado, prioridad, categoria, None,
                               departamento, fecha_desde, fecha_hasta, asignado_a_id)
    abiertos = [t for t in todos if t["estado"] in ("abierto", "en_progreso")]
    cerrados = [t for t in todos if t["estado"] in ("resuelto", "cerrado") and t.get("resuelto_en")]

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(f"Reporte de tickets — {empresa['nombre'] if empresa else ''}", styles["Title"]))
    if fecha_desde or fecha_hasta:
        rango = f"Del {fecha_desde or '…'} al {fecha_hasta or '…'}"
        elementos.append(Paragraph(rango, styles["Normal"]))
    elementos.append(Paragraph(f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 16))

    elementos.append(Paragraph(f"Tickets abiertos ({len(abiertos)})", styles["Heading2"]))
    datos_abiertos = [["Folio", "Departamento", "Solicitante", "Prioridad", "Estado", "Creado"]]
    for t in abiertos:
        datos_abiertos.append([
            t["folio"], t["departamento"], t["solicitante_nombre"],
            t["prioridad"], t["estado"], t["creado_en"][:16].replace("T", " "),
        ])
    tabla1 = Table(datos_abiertos, repeatRows=1)
    tabla1.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D8192F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elementos.append(tabla1)
    elementos.append(Spacer(1, 20))

    elementos.append(Paragraph(f"Tickets cerrados y tiempo de resolución ({len(cerrados)})", styles["Heading2"]))
    datos_cerrados = [["Folio", "Departamento", "Solicitante", "Creado", "Cerrado", "Horas"]]
    for t in cerrados:
        creado = dt.fromisoformat(t["creado_en"])
        resuelto = dt.fromisoformat(t["resuelto_en"])
        horas = round((resuelto - creado).total_seconds() / 3600, 1)
        datos_cerrados.append([
            t["folio"], t["departamento"], t["solicitante_nombre"],
            t["creado_en"][:16].replace("T", " "), t["resuelto_en"][:16].replace("T", " "), str(horas),
        ])
    tabla2 = Table(datos_cerrados, repeatRows=1)
    tabla2.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#74767A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elementos.append(tabla2)

    doc.build(elementos)
    buffer.seek(0)

    nombre_archivo = f"reporte_tickets_{dt.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.read(), media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


@app.get("/api/tickets/{ticket_id}")
def api_detalle_ticket(ticket_id: int, usuario: dict = Depends(requiere_ver_tickets)):
    ticket = db.obtener_ticket(ticket_id, empresa_id=usuario["empresa_id"])
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    if not _puede_ver_ticket(usuario, ticket):
        raise HTTPException(status_code=403, detail="No puedes ver este ticket")
    return ticket


@app.post("/api/tickets")
def api_crear_ticket(payload: NuevoTicket, usuario: dict = Depends(requiere_ver_tickets)):
    departamentos_validos = {d["nombre"] for d in db.listar_departamentos(usuario["empresa_id"])}
    categorias_validas = {c["nombre"] for c in db.listar_categorias(usuario["empresa_id"])}
    if payload.departamento not in departamentos_validos:
        raise HTTPException(status_code=400, detail="Departamento inválido")
    if payload.categoria not in categorias_validas:
        raise HTTPException(status_code=400, detail="Categoría inválida")
    if payload.prioridad not in db.PRIORIDADES:
        raise HTTPException(status_code=400, detail="Prioridad inválida")

    ticket = db.crear_ticket(usuario["empresa_id"], payload.departamento, payload.descripcion, payload.categoria, payload.prioridad, usuario["id"])
    tecnicos = db.listar_tecnicos_activos(usuario["empresa_id"])
    notifications.notificar_nuevo_ticket(tecnicos, ticket)
    if ticket.get("asignado_a_id"):
        tecnico_asignado = next((t for t in tecnicos if t["id"] == ticket["asignado_a_id"]), None)
        if tecnico_asignado:
            notifications.notificar_asignacion(tecnico_asignado, ticket)
    return ticket


@app.patch("/api/tickets/{ticket_id}")
def api_actualizar_ticket(ticket_id: int, payload: ActualizacionTicket, usuario: dict = Depends(requiere_staff)):
    usuario = _con_permisos(usuario)
    ticket_antes = db.obtener_ticket(ticket_id, empresa_id=usuario["empresa_id"])
    if not ticket_antes:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    if not _puede_ver_ticket(usuario, ticket_antes):
        raise HTTPException(status_code=403, detail="No puedes modificar este ticket")
    if payload.estado and payload.estado not in db.ESTADOS:
        raise HTTPException(status_code=400, detail="Estado inválido")
    if payload.prioridad and payload.prioridad not in db.PRIORIDADES:
        raise HTTPException(status_code=400, detail="Prioridad inválida")

    ticket = db.actualizar_ticket(ticket_id, payload.estado, payload.prioridad, payload.asignado_a_id)

    if payload.asignado_a_id and payload.asignado_a_id != ticket_antes.get("asignado_a_id"):
        tecnico = next((t for t in db.listar_tecnicos_activos(usuario["empresa_id"]) if t["id"] == payload.asignado_a_id), None)
        if tecnico:
            notifications.notificar_asignacion(tecnico, ticket)

    return ticket


@app.delete("/api/tickets/{ticket_id}")
def api_eliminar_ticket(ticket_id: int, usuario: dict = Depends(requiere_admin_completo)):
    if not db.eliminar_ticket(usuario["empresa_id"], ticket_id):
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    return {"ok": True}


MAX_ADJUNTO_BASE64 = 7_000_000  # ~5MB de archivo real (base64 pesa ~33% más)


@app.post("/api/tickets/{ticket_id}/comentarios")
def api_comentar(ticket_id: int, payload: NuevoComentario, usuario: dict = Depends(requiere_empresa)):
    usuario = _con_permisos(usuario)
    ticket = db.obtener_ticket(ticket_id, empresa_id=usuario["empresa_id"])
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    if not _puede_ver_ticket(usuario, ticket):
        raise HTTPException(status_code=403, detail="No puedes comentar este ticket")
    if payload.archivo_base64 and len(payload.archivo_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="El archivo pesa demasiado (máximo 5MB)")
    return db.agregar_comentario(ticket_id, usuario["id"], payload.texto,
                                  payload.archivo_base64, payload.archivo_nombre, payload.archivo_tipo)


@app.post("/api/tickets/{ticket_id}/firmar")
def api_firmar(ticket_id: int, payload: NuevaFirma, usuario: dict = Depends(requiere_staff)):
    usuario = _con_permisos(usuario)
    ticket = db.obtener_ticket(ticket_id, empresa_id=usuario["empresa_id"])
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket no encontrado")
    if not _puede_ver_ticket(usuario, ticket):
        raise HTTPException(status_code=403, detail="No puedes firmar este ticket")
    if ticket["estado"] == "cerrado":
        raise HTTPException(status_code=400, detail="Este ticket ya está cerrado")
    return db.firmar_ticket(ticket_id, payload.firma, payload.firmado_por.strip())


@app.get("/api/stats")
def api_stats(usuario: dict = Depends(requiere_staff)):
    usuario = _con_permisos(usuario)
    asignado_a_id = usuario["id"] if usuario["rol"] == "tecnico" else None
    categoria = usuario.get("restriccion_categoria") if usuario["rol"] == "admin" else None
    return db.estadisticas(usuario["empresa_id"], asignado_a_id, categoria)


# ==================== EQUIPOS (inventario) ====================

class NuevoEquipo(BaseModel):
    tipo: str
    nombre: str = Field(min_length=1, max_length=120)
    marca: Optional[str] = None
    modelo: Optional[str] = None
    numero_serie: Optional[str] = None
    departamento: Optional[str] = None
    responsable: Optional[str] = None
    fecha_adquisicion: Optional[str] = None
    notas: Optional[str] = None
    sucursal_id: Optional[int] = None
    usuario_id: Optional[int] = None
    usuario_microsip: Optional[str] = None
    password_microsip: Optional[str] = None


class ActualizacionEquipo(BaseModel):
    tipo: Optional[str] = None
    nombre: Optional[str] = None
    marca: Optional[str] = None
    modelo: Optional[str] = None
    numero_serie: Optional[str] = None
    departamento: Optional[str] = None
    responsable: Optional[str] = None
    estado: Optional[str] = None
    fecha_adquisicion: Optional[str] = None
    notas: Optional[str] = None
    sucursal_id: Optional[int] = None
    usuario_id: Optional[int] = None
    usuario_microsip: Optional[str] = None
    password_microsip: Optional[str] = None


CAMPOS_EQUIPO_SOLO_ADMIN = ["sucursal_id", "sucursal_nombre", "departamento", "usuario_id", "usuario_nombre", "usuario_microsip", "password_microsip"]


def _filtrar_equipo_por_rol(equipo, rol):
    if rol == "admin":
        return equipo
    return {k: v for k, v in equipo.items() if k not in CAMPOS_EQUIPO_SOLO_ADMIN}


@app.get("/api/equipos")
def api_listar_equipos(tipo: Optional[str] = None, estado: Optional[str] = None, usuario: dict = Depends(requiere_acceso_equipos)):
    equipos = db.listar_equipos(usuario["empresa_id"], tipo, estado)
    return [_filtrar_equipo_por_rol(e, usuario["rol"]) for e in equipos]


def _agrupar_equipos_por_departamento(equipos):
    grupos = {}
    for e in equipos:
        depto = e.get("departamento") or "Sin departamento"
        grupos.setdefault(depto, []).append(e)
    return dict(sorted(grupos.items()))


NOMBRES_TIPO_EQUIPO = {
    "computadora": "Computadora", "laptop": "Laptop", "monitor": "Monitor", "impresora": "Impresora",
    "escaner": "Escáner", "servidor": "Servidor",
    "mouse": "Mouse", "mouse_inalambrico": "Mouse inalámbrico", "teclado": "Teclado", "teclado_inalambrico": "Teclado inalámbrico",
    "router": "Router", "switch": "Switch", "modem": "Módem", "punto_acceso": "Punto de acceso (WiFi)",
    "dvr": "DVR", "camara_seguridad": "Cámara de seguridad", "no_break": "No-break (UPS)", "regulador": "Regulador de voltaje",
    "telefono": "Teléfono", "telefono_ip": "Teléfono IP", "proyector": "Proyector", "bocinas": "Bocinas", "microfono": "Micrófono",
    "tablet": "Tablet", "lector_codigo_barras": "Lector de código de barras", "disco_duro_externo": "Disco duro externo",
    "red": "Red", "otro": "Otro",
}
NOMBRES_ESTADO_EQUIPO = {
    "nuevo": "Nuevo", "buen_estado": "Buen estado", "sugerencia_cambio": "Sugerencia de cambio",
    "en_proceso_cambio": "En proceso de cambio", "cambio_urgente": "Cambio urgente", "baja": "Baja",
}


@app.get("/api/equipos/reporte.pdf")
def reporte_equipos_pdf(usuario: dict = Depends(requiere_acceso_equipos)):
    from io import BytesIO
    from datetime import datetime as dt

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    empresa = db.obtener_empresa(usuario["empresa_id"])
    equipos = db.listar_equipos(usuario["empresa_id"])
    grupos = _agrupar_equipos_por_departamento(equipos)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph(f"Inventario de equipos por departamento — {empresa['nombre'] if empresa else ''}", styles["Title"]))
    elementos.append(Paragraph(f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')} — {len(equipos)} equipos activos", styles["Normal"]))
    elementos.append(Spacer(1, 16))

    for depto, items in grupos.items():
        elementos.append(Paragraph(f"{depto} ({len(items)})", styles["Heading2"]))
        datos = [["Nombre", "Tipo", "Marca/Modelo", "N° Serie", "Responsable", "Estado"]]
        for e in items:
            datos.append([
                e["nombre"], NOMBRES_TIPO_EQUIPO.get(e["tipo"], e["tipo"]),
                " / ".join(filter(None, [e.get("marca"), e.get("modelo")])) or "—",
                e.get("numero_serie") or "—", e.get("responsable") or "—",
                NOMBRES_ESTADO_EQUIPO.get(e["estado"], e["estado"]),
            ])
        tabla = Table(datos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D8192F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        elementos.append(tabla)
        elementos.append(Spacer(1, 18))

    if not equipos:
        elementos.append(Paragraph("No hay equipos registrados en el inventario.", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)

    nombre_archivo = f"inventario_equipos_{dt.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.read(), media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


@app.get("/api/equipos/reporte.xlsx")
def reporte_equipos_xlsx(usuario: dict = Depends(requiere_acceso_equipos)):
    from io import BytesIO
    from datetime import datetime as dt

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    empresa = db.obtener_empresa(usuario["empresa_id"])
    equipos = db.listar_equipos(usuario["empresa_id"])
    grupos = _agrupar_equipos_por_departamento(equipos)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    encabezados = ["Departamento", "Nombre", "Tipo", "Marca", "Modelo", "N° Serie", "Responsable", "Estado", "Fecha adquisición", "Notas"]
    ws.append(encabezados)
    for col_idx, _ in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="D8192F", end_color="D8192F", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    fila = 2
    for depto, items in grupos.items():
        for e in items:
            ws.append([
                depto, e["nombre"], NOMBRES_TIPO_EQUIPO.get(e["tipo"], e["tipo"]),
                e.get("marca") or "", e.get("modelo") or "", e.get("numero_serie") or "",
                e.get("responsable") or "", NOMBRES_ESTADO_EQUIPO.get(e["estado"], e["estado"]),
                (e.get("fecha_adquisicion") or "")[:10], e.get("notas") or "",
            ])
            fila += 1

    for col_idx, encabezado in enumerate(encabezados, start=1):
        ancho = max(len(encabezado), 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho + 4

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nombre_archivo = f"inventario_equipos_{dt.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


@app.post("/api/equipos")
def api_crear_equipo(payload: NuevoEquipo, usuario: dict = Depends(requiere_acceso_equipos)):
    if payload.tipo not in db.TIPOS_EQUIPO:
        raise HTTPException(status_code=400, detail="Tipo de equipo inválido")
    if usuario["rol"] != "admin":
        payload.sucursal_id = None
        payload.departamento = None
        payload.usuario_id = None
        payload.usuario_microsip = None
        payload.password_microsip = None
    if payload.sucursal_id and not db.obtener_sucursal_reparacion(usuario["empresa_id"], payload.sucursal_id):
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    if payload.usuario_id and not any(u["id"] == payload.usuario_id for u in db.listar_usuarios(usuario["empresa_id"])):
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    equipo = db.crear_equipo(usuario["empresa_id"], payload.tipo, payload.nombre, payload.marca, payload.modelo,
                              payload.numero_serie, payload.departamento, payload.responsable,
                              payload.fecha_adquisicion, payload.notas, payload.sucursal_id,
                              payload.usuario_id, payload.usuario_microsip, payload.password_microsip)
    return _filtrar_equipo_por_rol(equipo, usuario["rol"])


@app.patch("/api/equipos/{equipo_id}")
def api_actualizar_equipo(equipo_id: int, payload: ActualizacionEquipo, usuario: dict = Depends(requiere_acceso_equipos)):
    if not db.obtener_equipo(usuario["empresa_id"], equipo_id):
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    if payload.tipo and payload.tipo not in db.TIPOS_EQUIPO:
        raise HTTPException(status_code=400, detail="Tipo de equipo inválido")
    if payload.estado and payload.estado not in db.ESTADOS_EQUIPO:
        raise HTTPException(status_code=400, detail="Estado de equipo inválido")
    datos = payload.dict(exclude_unset=True)
    if usuario["rol"] != "admin":
        for campo in ("sucursal_id", "departamento", "usuario_id", "usuario_microsip", "password_microsip"):
            datos.pop(campo, None)
    if datos.get("sucursal_id") and not db.obtener_sucursal_reparacion(usuario["empresa_id"], datos["sucursal_id"]):
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    if datos.get("usuario_id") and not any(u["id"] == datos["usuario_id"] for u in db.listar_usuarios(usuario["empresa_id"])):
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    equipo = db.actualizar_equipo(usuario["empresa_id"], equipo_id, **datos)
    return _filtrar_equipo_por_rol(equipo, usuario["rol"])


@app.get("/api/equipos/{equipo_id}")
def api_detalle_equipo(equipo_id: int, usuario: dict = Depends(requiere_empresa)):
    """Detalle de un solo equipo — cualquier persona de la empresa puede
    consultarlo (por ejemplo, para ver el texto de su propia responsiva antes
    de firmarla, sin necesitar acceso al módulo completo de Equipos)."""
    equipo = db.obtener_equipo(usuario["empresa_id"], equipo_id)
    if not equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    return equipo


@app.get("/api/equipos/{equipo_id}/carta-responsiva.pdf")
def api_carta_responsiva_equipo(equipo_id: int, usuario: dict = Depends(requiere_acceso_equipos)):
    equipo = db.obtener_equipo(usuario["empresa_id"], equipo_id)
    if not equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    empresa = db.obtener_empresa(usuario["empresa_id"])
    pdf_bytes = pdfs_equipos.generar_carta_responsiva(equipo, empresa)
    nombre_archivo = f"Carta_Responsiva_{(equipo.get('nombre') or 'equipo')[:30].replace(' ', '_')}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


class FirmaResponsivaEquipo(BaseModel):
    firma_base64: str = Field(min_length=100)


@app.post("/api/equipos/{equipo_id}/firmar-responsiva")
def api_firmar_responsiva_equipo(equipo_id: int, payload: FirmaResponsivaEquipo, usuario: dict = Depends(requiere_empresa)):
    """Cualquier persona puede firmar la responsiva del equipo que TIENE
    asignado a su nombre — no hace falta ser admin/técnico para esto, es su
    propio compromiso de resguardo. El administrador también puede firmarla
    en su nombre si hace falta (ej. la persona no tiene acceso al sistema)."""
    equipo = db.obtener_equipo(usuario["empresa_id"], equipo_id)
    if not equipo:
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    if usuario["rol"] != "admin" and equipo.get("usuario_id") != usuario["id"]:
        raise HTTPException(status_code=403, detail="Este equipo no está asignado a tu nombre")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    return db.firmar_responsiva_equipo(usuario["empresa_id"], equipo_id, payload.firma_base64)


@app.get("/api/mis-pendientes")
def api_mis_pendientes(usuario: dict = Depends(requiere_empresa)):
    """Todo lo que le falta resolver a esta persona en un solo lugar: equipos con
    responsiva sin firmar, proyectos esperando su firma de compromiso, y tareas de
    proyecto que tiene asignadas y sin terminar — para el apartado 'Mis tareas'."""
    equipos_pendientes = db.listar_equipos_pendientes_firma(usuario["empresa_id"], usuario["id"])
    proyectos_pendientes = db.listar_proyectos_pendientes_firma_usuario(usuario["empresa_id"], usuario["id"])
    tareas_pendientes = db.listar_tareas_proyecto_usuario(usuario["empresa_id"], usuario["id"])
    return {
        "equipos_pendientes": equipos_pendientes,
        "proyectos_pendientes": proyectos_pendientes,
        "tareas_pendientes": tareas_pendientes,
    }


@app.delete("/api/equipos/{equipo_id}")
def api_dar_de_baja_equipo(equipo_id: int, usuario: dict = Depends(requiere_acceso_equipos)):
    if not db.obtener_equipo(usuario["empresa_id"], equipo_id):
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    db.dar_de_baja_equipo(usuario["empresa_id"], equipo_id)
    return {"ok": True}


# ==================== MANTENIMIENTOS PROGRAMADOS ====================

class NuevoMantenimiento(BaseModel):
    equipo_id: int
    tipo: str = "preventivo"
    descripcion: str = Field(min_length=1)
    fecha_programada: str
    frecuencia: str = "unica"
    notas: Optional[str] = None
    tecnico_asignado_id: Optional[int] = None


class MarcarRealizado(BaseModel):
    realizado_por: str = Field(min_length=1, max_length=120)
    notas: Optional[str] = None


class ReprogramarMantenimiento(BaseModel):
    fecha_programada: Optional[str] = None
    descripcion: Optional[str] = None
    frecuencia: Optional[str] = None


@app.get("/api/mantenimientos")
def api_listar_mantenimientos(estado: Optional[str] = None, equipo_id: Optional[int] = None,
                               usuario: dict = Depends(requiere_acceso_equipos)):
    return db.listar_mantenimientos(usuario["empresa_id"], estado, equipo_id)


@app.post("/api/mantenimientos")
def api_crear_mantenimiento(payload: NuevoMantenimiento, usuario: dict = Depends(requiere_acceso_equipos)):
    if not db.obtener_equipo(usuario["empresa_id"], payload.equipo_id):
        raise HTTPException(status_code=404, detail="Equipo no encontrado")
    if payload.tipo not in db.TIPOS_MANTENIMIENTO:
        raise HTTPException(status_code=400, detail="Tipo de mantenimiento inválido")
    if payload.frecuencia not in db.FRECUENCIAS_MANTENIMIENTO:
        raise HTTPException(status_code=400, detail="Frecuencia inválida")
    mant_id = db.crear_mantenimiento(usuario["empresa_id"], payload.equipo_id, payload.tipo, payload.descripcion,
                                      payload.fecha_programada, payload.frecuencia, payload.notas,
                                      tecnico_asignado_id=payload.tecnico_asignado_id, creado_por_id=usuario["id"])
    return {"id": mant_id}


@app.post("/api/mantenimientos/{mantenimiento_id}/realizar")
def api_marcar_realizado(mantenimiento_id: int, payload: MarcarRealizado, usuario: dict = Depends(requiere_acceso_equipos)):
    resultado = db.marcar_mantenimiento_realizado(usuario["empresa_id"], mantenimiento_id, payload.realizado_por, payload.notas)
    if not resultado:
        raise HTTPException(status_code=404, detail="Mantenimiento no encontrado")
    return resultado


@app.patch("/api/mantenimientos/{mantenimiento_id}")
def api_reprogramar_mantenimiento(mantenimiento_id: int, payload: ReprogramarMantenimiento, usuario: dict = Depends(requiere_acceso_equipos)):
    if payload.frecuencia and payload.frecuencia not in db.FRECUENCIAS_MANTENIMIENTO:
        raise HTTPException(status_code=400, detail="Frecuencia inválida")
    db.reprogramar_mantenimiento(usuario["empresa_id"], mantenimiento_id, payload.fecha_programada,
                                  payload.descripcion, payload.frecuencia)
    return {"ok": True}


@app.delete("/api/mantenimientos/{mantenimiento_id}")
def api_eliminar_mantenimiento(mantenimiento_id: int, usuario: dict = Depends(requiere_acceso_equipos)):
    db.eliminar_mantenimiento(usuario["empresa_id"], mantenimiento_id)
    return {"ok": True}


# ==================== PROYECTOS ====================

class NuevaTareaProyecto(BaseModel):
    usuario_id: int
    descripcion: str = Field(min_length=1)
    fecha_limite: Optional[str] = None


class NuevoProyecto(BaseModel):
    nombre: str = Field(min_length=1, max_length=160)
    descripcion: Optional[str] = None
    fecha_estimada: Optional[str] = None
    participantes_usuarios: Optional[list[int]] = None
    participantes_departamentos: Optional[list[str]] = None
    tareas: Optional[list[NuevaTareaProyecto]] = None


class ActualizacionProyecto(BaseModel):
    nombre: Optional[str] = None
    descripcion: Optional[str] = None
    fecha_estimada: Optional[str] = None


class CambioEstadoProyecto(BaseModel):
    estado: str


class NuevoParticipanteUsuario(BaseModel):
    usuario_id: int


class NuevoParticipanteDepartamento(BaseModel):
    departamento: str


class NuevaActualizacionProyecto(BaseModel):
    texto: str = Field(min_length=1)
    archivo_base64: Optional[str] = None
    archivo_nombre: Optional[str] = None
    archivo_tipo: Optional[str] = None


def _puede_ver_proyecto(usuario, proyecto):
    if usuario["rol"] == "admin":
        return True
    return any(p["id"] == usuario["id"] for p in proyecto["participantes_usuarios"])


@app.get("/api/proyectos")
def api_listar_proyectos(estado: Optional[str] = None, usuario: dict = Depends(requiere_empresa)):
    participante_id = usuario["id"] if usuario["rol"] in ("usuario", "tecnico") else None
    return db.listar_proyectos(usuario["empresa_id"], participante_id, estado)


@app.post("/api/proyectos")
def api_crear_proyecto(payload: NuevoProyecto, usuario: dict = Depends(requiere_staff)):
    participantes_usuarios = list(payload.participantes_usuarios or [])
    if usuario["rol"] == "tecnico" and usuario["id"] not in participantes_usuarios:
        participantes_usuarios.append(usuario["id"])  # para que quien lo crea siempre lo pueda ver después
    if payload.tareas:
        ids_invalidos = [t.usuario_id for t in payload.tareas if t.usuario_id not in participantes_usuarios]
        if ids_invalidos:
            raise HTTPException(status_code=400, detail="No puedes asignar una tarea a alguien que no es participante del proyecto")
    tareas = [t.dict() for t in payload.tareas] if payload.tareas else None
    proyecto_id = db.crear_proyecto(
        usuario["empresa_id"], payload.nombre, payload.descripcion, payload.fecha_estimada, usuario["id"],
        participantes_usuarios, payload.participantes_departamentos, tareas,
    )
    return {"id": proyecto_id}


@app.get("/api/proyectos/reporte.pdf")
def reporte_proyectos_pdf(usuario: dict = Depends(requiere_empresa)):
    from io import BytesIO
    from datetime import datetime as dt

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    empresa = db.obtener_empresa(usuario["empresa_id"])
    participante_id = usuario["id"] if usuario["rol"] == "usuario" else None
    proyectos = db.listar_proyectos(usuario["empresa_id"], participante_id)

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Proyectos — {empresa['nombre'] if empresa else ''}", styles["Title"]),
        Paragraph(f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')} — {len(proyectos)} proyecto(s)", styles["Normal"]),
        Spacer(1, 16),
    ]

    datos = [["Nombre", "Estado", "Participantes", "Fecha estimada", "Días transcurridos", "Creado por"]]
    for p in proyectos:
        participantes = ", ".join(
            [u["nombre_completo"] for u in p["participantes_usuarios"]] + p["participantes_departamentos"]
        ) or "—"
        datos.append([
            p["nombre"], NOMBRES_ESTADO_PROYECTO_PDF.get(p["estado"], p["estado"]), participantes,
            (p.get("fecha_estimada") or "—")[:10],
            str(p["dias_transcurridos"]) if p.get("dias_transcurridos") is not None else "—",
            p["creado_por_nombre"],
        ])
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D8192F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elementos.append(tabla)
    if not proyectos:
        elementos.append(Paragraph("No hay proyectos registrados.", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    nombre_archivo = f"proyectos_{dt.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.read(), media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


@app.get("/api/proyectos/reporte.xlsx")
def reporte_proyectos_xlsx(usuario: dict = Depends(requiere_empresa)):
    from io import BytesIO
    from datetime import datetime as dt

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    participante_id = usuario["id"] if usuario["rol"] == "usuario" else None
    proyectos = db.listar_proyectos(usuario["empresa_id"], participante_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    encabezados = ["Nombre", "Descripción", "Estado", "Personas", "Departamentos", "Fecha estimada",
                   "Fecha inicio", "Fecha completado", "Días transcurridos", "Creado por"]
    ws.append(encabezados)
    for col_idx, _ in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="D8192F", end_color="D8192F", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    for p in proyectos:
        ws.append([
            p["nombre"], p.get("descripcion") or "", NOMBRES_ESTADO_PROYECTO_PDF.get(p["estado"], p["estado"]),
            ", ".join(u["nombre_completo"] for u in p["participantes_usuarios"]),
            ", ".join(p["participantes_departamentos"]),
            (p.get("fecha_estimada") or "")[:10], (p.get("fecha_inicio") or "")[:10],
            (p.get("fecha_completado") or "")[:10],
            p["dias_transcurridos"] if p.get("dias_transcurridos") is not None else "",
            p["creado_por_nombre"],
        ])

    for col_idx, encabezado in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(encabezado), 14) + 4
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"proyectos_{dt.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


@app.get("/api/proyectos/{proyecto_id}")
def api_detalle_proyecto(proyecto_id: int, usuario: dict = Depends(requiere_empresa)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    return proyecto


@app.post("/api/proyectos/{proyecto_id}/tareas")
def api_crear_tarea_proyecto(proyecto_id: int, payload: NuevaTareaProyecto, usuario: dict = Depends(requiere_staff)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not any(p["id"] == payload.usuario_id for p in proyecto["participantes_usuarios"]):
        raise HTTPException(status_code=400, detail="Esa persona no es participante del proyecto — agrégala primero")
    db.crear_tarea_proyecto(proyecto_id, payload.usuario_id, payload.descripcion, payload.fecha_limite)
    return db.obtener_proyecto(usuario["empresa_id"], proyecto_id)


class CambioEstadoTareaProyecto(BaseModel):
    estado: str


@app.patch("/api/proyectos/{proyecto_id}/tareas/{tarea_id}")
def api_cambiar_estado_tarea_proyecto(proyecto_id: int, tarea_id: int, payload: CambioEstadoTareaProyecto, usuario: dict = Depends(requiere_empresa)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    tarea = db.obtener_tarea_proyecto(tarea_id)
    if not tarea or tarea["proyecto_id"] != proyecto_id:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    if payload.estado not in db.ESTADOS_TAREA_PROYECTO:
        raise HTTPException(status_code=400, detail="Estado inválido")
    # Solo quien tiene la tarea asignada, o el staff, puede cambiar su estado
    if usuario["rol"] not in ("admin", "tecnico") and usuario["id"] != tarea["usuario_id"]:
        raise HTTPException(status_code=403, detail="Esta tarea no te pertenece")
    db.cambiar_estado_tarea_proyecto(tarea_id, payload.estado)
    return db.obtener_proyecto(usuario["empresa_id"], proyecto_id)


@app.delete("/api/proyectos/{proyecto_id}/tareas/{tarea_id}")
def api_eliminar_tarea_proyecto(proyecto_id: int, tarea_id: int, usuario: dict = Depends(requiere_staff)):
    tarea = db.obtener_tarea_proyecto(tarea_id)
    if not tarea or tarea["proyecto_id"] != proyecto_id:
        raise HTTPException(status_code=404, detail="Tarea no encontrada")
    db.eliminar_tarea_proyecto(tarea_id)
    return {"ok": True}


@app.patch("/api/proyectos/{proyecto_id}")
def api_actualizar_proyecto(proyecto_id: int, payload: ActualizacionProyecto, usuario: dict = Depends(requiere_staff)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    db.actualizar_proyecto(usuario["empresa_id"], proyecto_id, payload.nombre, payload.descripcion, payload.fecha_estimada)
    return {"ok": True}


@app.post("/api/proyectos/{proyecto_id}/iniciar")
def api_iniciar_proyecto(proyecto_id: int, usuario: dict = Depends(requiere_staff)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    resultado = db.iniciar_proyecto(usuario["empresa_id"], proyecto_id)
    if not resultado["ok"]:
        nombres = ", ".join(p["nombre_completo"] for p in resultado["pendientes"])
        raise HTTPException(status_code=400, detail=f"Todavía falta que firmen o digan por qué no están conformes: {nombres}")
    return {"ok": True}


class FirmaParticipanteProyecto(BaseModel):
    firma_base64: str = Field(min_length=100)


class NoConformeProyecto(BaseModel):
    motivo: str = Field(min_length=3)


@app.post("/api/proyectos/{proyecto_id}/firmar-participante")
def api_firmar_participante_proyecto(proyecto_id: int, payload: FirmaParticipanteProyecto, usuario: dict = Depends(requiere_empresa)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not any(p["id"] == usuario["id"] for p in proyecto["participantes_usuarios"]):
        raise HTTPException(status_code=403, detail="No eres participante de este proyecto")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    db.firmar_participante_proyecto(proyecto_id, usuario["id"], payload.firma_base64)
    return db.obtener_proyecto(usuario["empresa_id"], proyecto_id)


@app.post("/api/proyectos/{proyecto_id}/no-conforme")
def api_no_conforme_proyecto(proyecto_id: int, payload: NoConformeProyecto, usuario: dict = Depends(requiere_empresa)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not any(p["id"] == usuario["id"] for p in proyecto["participantes_usuarios"]):
        raise HTTPException(status_code=403, detail="No eres participante de este proyecto")
    db.marcar_no_conforme_proyecto(proyecto_id, usuario["id"], payload.motivo.strip())
    return db.obtener_proyecto(usuario["empresa_id"], proyecto_id)


@app.patch("/api/proyectos/{proyecto_id}/estado")
def api_cambiar_estado_proyecto(proyecto_id: int, payload: CambioEstadoProyecto, usuario: dict = Depends(requiere_staff)):
    if payload.estado not in db.ESTADOS_PROYECTO:
        raise HTTPException(status_code=400, detail="Estado inválido")
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    db.cambiar_estado_proyecto(usuario["empresa_id"], proyecto_id, payload.estado)
    return {"ok": True}


@app.post("/api/proyectos/{proyecto_id}/participantes/usuarios")
def api_agregar_participante_usuario(proyecto_id: int, payload: NuevoParticipanteUsuario, usuario: dict = Depends(requiere_staff)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    db.agregar_participante_usuario(proyecto_id, payload.usuario_id)
    return {"ok": True}


@app.delete("/api/proyectos/{proyecto_id}/participantes/usuarios/{usuario_id}")
def api_quitar_participante_usuario(proyecto_id: int, usuario_id: int, usuario: dict = Depends(requiere_staff)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    db.quitar_participante_usuario(proyecto_id, usuario_id)
    return {"ok": True}


@app.post("/api/proyectos/{proyecto_id}/participantes/departamentos")
def api_agregar_participante_departamento(proyecto_id: int, payload: NuevoParticipanteDepartamento, usuario: dict = Depends(requiere_staff)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    db.agregar_participante_departamento(proyecto_id, payload.departamento)
    return {"ok": True}


@app.delete("/api/proyectos/{proyecto_id}/participantes/departamentos/{departamento}")
def api_quitar_participante_departamento(proyecto_id: int, departamento: str, usuario: dict = Depends(requiere_staff)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    db.quitar_participante_departamento(proyecto_id, departamento)
    return {"ok": True}


@app.post("/api/proyectos/{proyecto_id}/actualizaciones")
def api_comentar_proyecto(proyecto_id: int, payload: NuevaActualizacionProyecto, usuario: dict = Depends(requiere_empresa)):
    proyecto = db.obtener_proyecto(usuario["empresa_id"], proyecto_id)
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    if not _puede_ver_proyecto(usuario, proyecto):
        raise HTTPException(status_code=403, detail="No participas en este proyecto")
    if payload.archivo_base64 and len(payload.archivo_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="El archivo pesa demasiado (máximo 5MB)")
    db.agregar_actualizacion_proyecto(proyecto_id, usuario["id"], payload.texto,
                                       payload.archivo_base64, payload.archivo_nombre, payload.archivo_tipo)
    return db.obtener_proyecto(usuario["empresa_id"], proyecto_id)


NOMBRES_ESTADO_PROYECTO_PDF = {
    "planificacion": "Planificación", "en_progreso": "En progreso", "pausado": "Pausado",
    "completado": "Completado", "cancelado": "Cancelado",
}


# ==================== COMPRAS ====================

class NuevoArticuloCompra(BaseModel):
    nombre: str = Field(min_length=1, max_length=160)
    proveedor: Optional[str] = None
    marca: Optional[str] = None
    foto_base64: Optional[str] = None
    notas: Optional[str] = None
    categoria: Optional[str] = None
    precio_unitario: Optional[float] = None
    stock_actual: Optional[int] = None
    stock_minimo: Optional[int] = None


class ActualizacionArticuloCompra(BaseModel):
    nombre: Optional[str] = None
    proveedor: Optional[str] = None
    marca: Optional[str] = None
    foto_base64: Optional[str] = None
    notas: Optional[str] = None
    categoria: Optional[str] = None
    precio_unitario: Optional[float] = None
    stock_minimo: Optional[int] = None


@app.get("/api/compras/articulos")
def api_listar_articulos_compra(usuario: dict = Depends(requiere_ver_compras)):
    return db.listar_articulos_compra(usuario["empresa_id"])


@app.get("/api/compras/categorias")
def api_listar_categorias_compra(usuario: dict = Depends(requiere_ver_compras)):
    return db.listar_categorias_compra(usuario["empresa_id"])


@app.post("/api/compras/articulos/catalogo-inicial")
def api_sembrar_catalogo_compras(usuario: dict = Depends(requiere_admin_compras)):
    """Carga de un solo golpe un catálogo de productos comunes (papelería, limpieza,
    ferretería, equipo de cómputo, cafetería, equipo de oficina) con precios de
    referencia — no duplica artículos que ya existan por nombre."""
    agregados = db.sembrar_catalogo_compras(usuario["empresa_id"])
    return {"agregados": agregados, "total_catalogo": len(db.CATALOGO_INICIAL_COMPRAS)}


@app.post("/api/compras/articulos")
def api_crear_articulo_compra(payload: NuevoArticuloCompra, usuario: dict = Depends(requiere_admin_compras)):
    if payload.foto_base64 and len(payload.foto_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La foto pesa demasiado (máximo 5MB)")
    articulo_id = db.crear_articulo_compra(usuario["empresa_id"], payload.nombre, payload.proveedor,
                                            payload.marca, payload.foto_base64, payload.notas, payload.categoria,
                                            payload.precio_unitario, payload.stock_actual or 0, payload.stock_minimo or 0)
    return {"id": articulo_id}


@app.patch("/api/compras/articulos/{articulo_id}")
def api_actualizar_articulo_compra(articulo_id: int, payload: ActualizacionArticuloCompra, usuario: dict = Depends(requiere_admin_compras)):
    if not db.obtener_articulo_compra(usuario["empresa_id"], articulo_id):
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    if payload.foto_base64 and len(payload.foto_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La foto pesa demasiado (máximo 5MB)")
    db.actualizar_articulo_compra(usuario["empresa_id"], articulo_id, **payload.dict(exclude_unset=True))
    return {"ok": True}


@app.delete("/api/compras/articulos/{articulo_id}")
def api_dar_de_baja_articulo_compra(articulo_id: int, usuario: dict = Depends(requiere_admin_compras)):
    if not db.obtener_articulo_compra(usuario["empresa_id"], articulo_id):
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    db.dar_de_baja_articulo_compra(usuario["empresa_id"], articulo_id)
    return {"ok": True}


class AjusteStockArticulo(BaseModel):
    delta: int  # positivo para sumar existencias, negativo para restar


@app.post("/api/compras/articulos/{articulo_id}/ajustar-stock")
def api_ajustar_stock_articulo(articulo_id: int, payload: AjusteStockArticulo, usuario: dict = Depends(requiere_admin_compras)):
    if not db.obtener_articulo_compra(usuario["empresa_id"], articulo_id):
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    nuevo_stock = db.ajustar_stock_articulo(usuario["empresa_id"], articulo_id, payload.delta)
    return {"stock_actual": nuevo_stock}


class NuevoCicloCompra(BaseModel):
    nombre: str = Field(min_length=1, max_length=160)
    frecuencia: str = "unica"
    fecha_programada: str
    categoria: Optional[str] = None


class NuevoPedidoCompra(BaseModel):
    articulo_id: Optional[int] = None
    articulo_libre: Optional[str] = None
    cantidad: int = Field(default=1, ge=1)
    sucursal_id: Optional[int] = None
    notas: Optional[str] = None


@app.get("/api/compras/ciclos")
def api_listar_ciclos_compra(estado: Optional[str] = None, usuario: dict = Depends(requiere_ver_compras)):
    return db.listar_ciclos_compra(usuario["empresa_id"], estado)


@app.post("/api/compras/ciclos")
def api_crear_ciclo_compra(payload: NuevoCicloCompra, usuario: dict = Depends(requiere_admin_compras)):
    if payload.frecuencia not in db.FRECUENCIAS_COMPRA:
        raise HTTPException(status_code=400, detail="Frecuencia inválida")
    ciclo_id = db.crear_ciclo_compra(usuario["empresa_id"], payload.nombre, payload.frecuencia,
                                      payload.fecha_programada, usuario["id"], payload.categoria)
    return {"id": ciclo_id}


@app.get("/api/compras/ciclos/{ciclo_id}")
def api_detalle_ciclo_compra(ciclo_id: int, usuario: dict = Depends(requiere_ver_compras)):
    ciclo = db.obtener_ciclo_compra(usuario["empresa_id"], ciclo_id)
    if not ciclo:
        raise HTTPException(status_code=404, detail="Ciclo no encontrado")
    if usuario["rol"] == "usuario":
        # Un empleado solo ve los pedidos de SU propia sucursal, nunca los de otras.
        # Si no tiene sucursal asignada, ve solo sus propios pedidos (no todos los "sin sucursal").
        mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
        if mi_sucursal_id:
            ciclo["pedidos"] = [p for p in ciclo["pedidos"] if p["sucursal_id"] == mi_sucursal_id]
        else:
            ciclo["pedidos"] = [p for p in ciclo["pedidos"] if p["usuario_id"] == usuario["id"]]
    return ciclo


@app.post("/api/compras/ciclos/{ciclo_id}/abrir")
def api_abrir_ciclo_compra(ciclo_id: int, usuario: dict = Depends(requiere_acceso_compras)):
    if not db.obtener_ciclo_compra(usuario["empresa_id"], ciclo_id):
        raise HTTPException(status_code=404, detail="Ciclo no encontrado")
    db.abrir_ciclo_compra(usuario["empresa_id"], ciclo_id)
    return {"ok": True}


@app.post("/api/compras/ciclos/{ciclo_id}/cerrar")
def api_cerrar_ciclo_compra(ciclo_id: int, usuario: dict = Depends(requiere_acceso_compras)):
    resultado = db.cerrar_ciclo_compra(usuario["empresa_id"], ciclo_id)
    if not resultado:
        raise HTTPException(status_code=404, detail="Ciclo no encontrado")
    ciclo_cerrado = db.obtener_ciclo_compra(usuario["empresa_id"], ciclo_id)
    masters = db.listar_usuarios_master(usuario["empresa_id"])
    if masters:
        notifications.notificar_ciclo_pendiente_autorizacion(masters, ciclo_cerrado, ciclo_cerrado["total_general"])
    return resultado


@app.post("/api/compras/ciclos/{ciclo_id}/pedidos")
def api_agregar_pedido_compra(ciclo_id: int, payload: NuevoPedidoCompra, usuario: dict = Depends(requiere_ver_compras)):
    ciclo = db.obtener_ciclo_compra(usuario["empresa_id"], ciclo_id)
    if not ciclo:
        raise HTTPException(status_code=404, detail="Ciclo no encontrado")
    if ciclo["estado"] in ("esperando_autorizacion", "cerrado"):
        raise HTTPException(status_code=400, detail="Este ciclo ya se marcó como surtido — ya no se pueden agregar pedidos")
    if not payload.articulo_id and not (payload.articulo_libre and payload.articulo_libre.strip()):
        raise HTTPException(status_code=400, detail="Elige un artículo del catálogo o escribe uno libre")
    if payload.articulo_id and not db.obtener_articulo_compra(usuario["empresa_id"], payload.articulo_id):
        raise HTTPException(status_code=404, detail="Artículo no encontrado")
    if payload.sucursal_id and not db.obtener_sucursal_reparacion(usuario["empresa_id"], payload.sucursal_id):
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    db.agregar_pedido_compra(ciclo_id, payload.articulo_id, usuario["id"], payload.cantidad,
                              payload.sucursal_id, payload.notas,
                              payload.articulo_libre.strip() if payload.articulo_libre else None)
    return db.obtener_ciclo_compra(usuario["empresa_id"], ciclo_id)


@app.delete("/api/compras/pedidos/{pedido_id}")
def api_eliminar_pedido_compra(pedido_id: int, usuario: dict = Depends(requiere_ver_compras)):
    es_staff = usuario["rol"] in ("admin", "tecnico")
    db.eliminar_pedido_compra(pedido_id, usuario["id"], es_staff)
    return {"ok": True}


@app.post("/api/compras/pedidos/{pedido_id}/listo")
def api_marcar_pedido_listo(pedido_id: int, usuario: dict = Depends(requiere_acceso_compras)):
    pedido = db.obtener_pedido_compra(usuario["empresa_id"], pedido_id)
    if not pedido:
        raise HTTPException(status_code=404, detail="Pedido no encontrado")
    db.marcar_pedido_listo(pedido_id)
    notifications.notificar_pedido_listo(
        {"telefono_whatsapp": pedido.get("usuario_telefono")}, pedido["articulo_nombre"], pedido["cantidad"],
    )
    ciclo = db.obtener_ciclo_compra(usuario["empresa_id"], pedido["ciclo_id"])
    return ciclo


def requiere_autorizacion_compras(usuario: dict = Depends(requiere_empresa_o_master)) -> dict:
    """Autorizar una compra ya cerrada es exclusivo del usuario master (para eso
    existe) y del administrador, como respaldo."""
    if usuario["rol"] not in ("admin", "master"):
        raise HTTPException(status_code=403, detail="No tienes permiso para autorizar compras")
    return usuario


class FirmaAutorizacionCompra(BaseModel):
    firma_base64: str = Field(min_length=100)


@app.get("/api/compras/autorizaciones")
def api_listar_autorizaciones_compra(usuario: dict = Depends(requiere_autorizacion_compras)):
    return db.listar_ciclos_pendientes_autorizacion(usuario["empresa_id"])


@app.post("/api/compras/ciclos/{ciclo_id}/autorizar")
def api_autorizar_ciclo_compra(ciclo_id: int, payload: FirmaAutorizacionCompra, usuario: dict = Depends(requiere_autorizacion_compras)):
    ciclo = db.obtener_ciclo_compra(usuario["empresa_id"], ciclo_id)
    if not ciclo:
        raise HTTPException(status_code=404, detail="Ciclo no encontrado")
    if ciclo["estado"] != "esperando_autorizacion":
        raise HTTPException(status_code=400, detail="Este ciclo no está esperando autorización (o ya fue autorizado)")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    db.autorizar_ciclo_compra(usuario["empresa_id"], ciclo_id, usuario["id"], payload.firma_base64)
    return db.obtener_ciclo_compra(usuario["empresa_id"], ciclo_id)


NOMBRES_FRECUENCIA_COMPRA_PDF = {"unica": "Única vez", "semanal": "Semanal", "quincenal": "Quincenal", "mensual": "Mensual"}
NOMBRES_ESTADO_CICLO_PDF = {"pendiente": "Pendiente", "abierto": "Abierto", "cerrado": "Cerrado"}


@app.get("/api/compras/reporte.pdf")
def reporte_compras_pdf(usuario: dict = Depends(requiere_ver_compras)):
    from io import BytesIO
    from datetime import datetime as dt

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    empresa = db.obtener_empresa(usuario["empresa_id"])
    ciclos = db.listar_ciclos_compra(usuario["empresa_id"])

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Compras — {empresa['nombre'] if empresa else ''}", styles["Title"]),
        Paragraph(f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')} — {len(ciclos)} ciclo(s)", styles["Normal"]),
        Spacer(1, 16),
    ]

    for c in ciclos:
        detalle = db.obtener_ciclo_compra(usuario["empresa_id"], c["id"])
        elementos.append(Paragraph(
            f"{c['nombre']} — {NOMBRES_FRECUENCIA_COMPRA_PDF.get(c['frecuencia'], c['frecuencia'])} — "
            f"{NOMBRES_ESTADO_CICLO_PDF.get(c['estado'], c['estado'])} ({c['fecha_programada'][:10]})",
            styles["Heading2"],
        ))
        datos = [["Artículo", "Cantidad", "Sucursal", "Pedido por", "Notas"]]
        for p in detalle["pedidos"]:
            datos.append([p["articulo_nombre"], str(p["cantidad"]), p.get("sucursal_nombre") or "—",
                          p["usuario_nombre"], p.get("notas") or "—"])
        if len(datos) == 1:
            datos.append(["— sin pedidos —", "", "", "", ""])
        tabla = Table(datos, repeatRows=1)
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D8192F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        elementos.append(tabla)
        elementos.append(Spacer(1, 18))

    if not ciclos:
        elementos.append(Paragraph("No hay ciclos de compra registrados.", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    nombre_archivo = f"compras_{dt.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.read(), media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


@app.get("/api/compras/reporte.xlsx")
def reporte_compras_xlsx(usuario: dict = Depends(requiere_ver_compras)):
    from io import BytesIO
    from datetime import datetime as dt

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    ciclos = db.listar_ciclos_compra(usuario["empresa_id"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"
    encabezados = ["Ciclo", "Frecuencia", "Estado del ciclo", "Fecha programada",
                   "Artículo", "Cantidad", "Sucursal", "Pedido por", "Notas"]
    ws.append(encabezados)
    for col_idx, _ in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="D8192F", end_color="D8192F", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    for c in ciclos:
        detalle = db.obtener_ciclo_compra(usuario["empresa_id"], c["id"])
        filas = detalle["pedidos"] or [None]
        for p in filas:
            ws.append([
                c["nombre"], NOMBRES_FRECUENCIA_COMPRA_PDF.get(c["frecuencia"], c["frecuencia"]),
                NOMBRES_ESTADO_CICLO_PDF.get(c["estado"], c["estado"]), c["fecha_programada"][:10],
                p["articulo_nombre"] if p else "", p["cantidad"] if p else "",
                (p.get("sucursal_nombre") or "") if p else "", p["usuario_nombre"] if p else "",
                (p.get("notas") or "") if p else "",
            ])

    for col_idx, encabezado in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(encabezado), 14) + 4
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"compras_{dt.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


# ==================== RECURSOS HUMANOS (incidencias) ====================

class NuevaIncidenciaRH(BaseModel):
    tipo: str
    fecha_inicio: str
    fecha_fin: Optional[str] = None
    motivo: Optional[str] = None
    foto_base64: Optional[str] = None
    horas: Optional[float] = None


class ResolverIncidenciaRH(BaseModel):
    estado: str  # 'aprobada' o 'rechazada'
    respuesta_admin: Optional[str] = None


@app.get("/api/rh/incidencias")
def api_listar_incidencias_rh(estado: Optional[str] = None, usuario: dict = Depends(requiere_ver_rh)):
    # El administrador ve las de todos; cualquier otro rol solo ve las suyas
    # (para que cada quien siga viendo el estatus de lo suyo, aunque siga
    # esperando al encargado de su sucursal). Cuando el administrador pide
    # "todas" sin filtro, no le mezclamos las que ni siquiera ha visto el
    # encargado todavía — esas están en la bandeja del encargado, no en la suya.
    usuario_id_filtro = None if usuario["rol"] == "admin" else usuario["id"]
    resultado = db.listar_incidencias_rh(usuario["empresa_id"], usuario_id_filtro, estado)
    if usuario["rol"] == "admin" and estado is None:
        resultado = [i for i in resultado if i["estado"] != "pendiente_encargado"]
    return resultado


@app.post("/api/rh/incidencias")
def api_crear_incidencia_rh(payload: NuevaIncidenciaRH, usuario: dict = Depends(requiere_ver_rh)):
    if payload.tipo not in db.TIPOS_INCIDENCIA_RH:
        raise HTTPException(status_code=400, detail="Tipo de incidencia inválido")
    if payload.foto_base64 and len(payload.foto_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La foto pesa demasiado (máximo 5MB)")
    if payload.horas is not None and payload.horas <= 0:
        raise HTTPException(status_code=400, detail="Las horas deben ser un número positivo")
    incidencia_id = db.crear_incidencia_rh(usuario["empresa_id"], usuario["id"], payload.tipo,
                                            payload.fecha_inicio, payload.fecha_fin, payload.motivo,
                                            payload.foto_base64, payload.horas)
    return {"id": incidencia_id}


def requiere_encargado_sucursal(usuario: dict = Depends(requiere_empresa)) -> dict:
    if usuario["rol"] != "encargado_sucursal":
        raise HTTPException(status_code=403, detail="Esta acción es solo para el encargado de sucursal")
    return usuario


@app.get("/api/rh/incidencias/pendientes-encargado")
def api_incidencias_pendientes_encargado(usuario: dict = Depends(requiere_encargado_sucursal)):
    mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
    if not mi_sucursal_id:
        return []
    return db.listar_incidencias_pendientes_encargado(usuario["empresa_id"], mi_sucursal_id)


class FirmaAceptacionEncargado(BaseModel):
    firma_base64: str = Field(min_length=100)


@app.post("/api/rh/incidencias/{incidencia_id}/aceptar-encargado")
def api_aceptar_incidencia_encargado(incidencia_id: int, payload: FirmaAceptacionEncargado, usuario: dict = Depends(requiere_encargado_sucursal)):
    """El encargado de sucursal firma para aceptar la incidencia de alguien de
    SU sucursal — recién ahí pasa a la bandeja de Recursos Humanos."""
    incidencia = db.obtener_incidencia_rh(usuario["empresa_id"], incidencia_id)
    if not incidencia:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
    sucursal_de_la_persona = db.obtener_sucursal_id_usuario(incidencia["usuario_id"])
    if not mi_sucursal_id or sucursal_de_la_persona != mi_sucursal_id:
        raise HTTPException(status_code=403, detail="Esta incidencia no es de tu sucursal")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    if not db.aceptar_incidencia_encargado(usuario["empresa_id"], incidencia_id, usuario["id"], payload.firma_base64):
        raise HTTPException(status_code=400, detail="Esta incidencia ya no está esperando tu firma (puede que ya se haya aceptado)")
    return {"ok": True}



@app.get("/api/rh/incidencias/{incidencia_id}")
def api_detalle_incidencia_rh(incidencia_id: int, usuario: dict = Depends(requiere_empresa)):
    incidencia = db.obtener_incidencia_rh(usuario["empresa_id"], incidencia_id)
    if not incidencia:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    if usuario["rol"] == "admin" or incidencia["usuario_id"] == usuario["id"]:
        return incidencia
    if usuario["rol"] == "encargado_sucursal":
        # También puede ver (para revisarla antes de firmar) cualquier
        # incidencia de alguien de SU MISMA sucursal, sin importar quién sea.
        mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
        sucursal_de_la_persona = db.obtener_sucursal_id_usuario(incidencia["usuario_id"])
        if mi_sucursal_id and sucursal_de_la_persona == mi_sucursal_id:
            return incidencia
    raise HTTPException(status_code=403, detail="No puedes ver la incidencia de alguien más")


@app.post("/api/rh/incidencias/{incidencia_id}/resolver")
def api_resolver_incidencia_rh(incidencia_id: int, payload: ResolverIncidenciaRH, usuario: dict = Depends(requiere_admin_rh)):
    if payload.estado not in ("aprobada", "rechazada", "pagada"):
        raise HTTPException(status_code=400, detail="Estado inválido")
    incidencia = db.obtener_incidencia_rh(usuario["empresa_id"], incidencia_id)
    if not incidencia:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    if incidencia["estado"] != "pendiente":
        raise HTTPException(status_code=400, detail="Esta incidencia ya fue resuelta")
    db.resolver_incidencia_rh(usuario["empresa_id"], incidencia_id, usuario["id"], payload.estado, payload.respuesta_admin)
    incidencia_resuelta = db.obtener_incidencia_rh(usuario["empresa_id"], incidencia_id)
    notifications.notificar_incidencia_rh_resuelta(
        {"telefono_whatsapp": incidencia_resuelta.get("usuario_telefono")}, incidencia_resuelta,
    )
    return incidencia_resuelta


class RespuestaPropuestaDiaSinGoce(BaseModel):
    acepta: bool


@app.post("/api/rh/incidencias/{incidencia_id}/responder-propuesta")
def api_responder_propuesta_dia_sin_goce(incidencia_id: int, payload: RespuestaPropuestaDiaSinGoce, usuario: dict = Depends(requiere_empresa)):
    """El empleado responde si acepta o no convertir sus 8 horas acumuladas
    en un día sin goce de sueldo. Si acepta, pasa a que su encargada de
    sucursal lo autorice (o directo a RH si no tiene encargada asignada)."""
    if not db.responder_propuesta_dia_sin_goce(usuario["empresa_id"], incidencia_id, usuario["id"], payload.acepta):
        raise HTTPException(status_code=400, detail="Esta propuesta ya no está esperando tu respuesta")
    return {"ok": True}


@app.delete("/api/rh/incidencias/{incidencia_id}")
def api_eliminar_incidencia_rh(incidencia_id: int, usuario: dict = Depends(requiere_empresa)):
    es_admin = usuario["rol"] == "admin"
    es_encargado_de_esa_persona = False
    if usuario["rol"] == "encargado_sucursal" and not es_admin:
        incidencia = db.obtener_incidencia_rh(usuario["empresa_id"], incidencia_id)
        if incidencia:
            mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
            sucursal_de_la_persona = db.obtener_sucursal_id_usuario(incidencia["usuario_id"])
            es_encargado_de_esa_persona = bool(mi_sucursal_id) and sucursal_de_la_persona == mi_sucursal_id
    ok = db.eliminar_incidencia_rh(usuario["empresa_id"], incidencia_id, usuario["id"], es_admin, es_encargado_de_esa_persona)
    if not ok:
        raise HTTPException(status_code=400, detail="No se pudo eliminar (no es tuya, no es de tu sucursal, o ya fue resuelta)")
    return {"ok": True}


class EdicionIncidenciaRH(BaseModel):
    tipo: Optional[str] = None
    fecha_inicio: Optional[str] = None
    fecha_fin: Optional[str] = None
    motivo: Optional[str] = None
    horas: Optional[float] = Field(default=None, ge=0, le=24)


@app.patch("/api/rh/incidencias/{incidencia_id}")
def api_editar_incidencia_rh(incidencia_id: int, payload: EdicionIncidenciaRH, usuario: dict = Depends(requiere_empresa)):
    """Para corregir una incidencia que se capturó mal — el administrador
    puede editar cualquiera; la encargada de sucursal, solo las de su gente
    y solo mientras sigan esperando su firma (no una vez que ya pasó a RH)."""
    incidencia = db.obtener_incidencia_rh(usuario["empresa_id"], incidencia_id)
    if not incidencia:
        raise HTTPException(status_code=404, detail="Incidencia no encontrada")
    if usuario["rol"] != "admin":
        if usuario["rol"] != "encargado_sucursal":
            raise HTTPException(status_code=403, detail="No tienes permiso para editar incidencias")
        mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
        sucursal_de_la_persona = db.obtener_sucursal_id_usuario(incidencia["usuario_id"])
        if not mi_sucursal_id or sucursal_de_la_persona != mi_sucursal_id:
            raise HTTPException(status_code=403, detail="Esta incidencia no es de tu sucursal")
        if incidencia["estado"] != "pendiente_encargado":
            raise HTTPException(status_code=400, detail="Ya no se puede editar — esta incidencia ya pasó a Recursos Humanos")
    if payload.tipo and payload.tipo not in db.TIPOS_INCIDENCIA_RH:
        raise HTTPException(status_code=400, detail="Tipo de incidencia inválido")
    return db.editar_incidencia_rh(usuario["empresa_id"], incidencia_id, payload.tipo, payload.fecha_inicio,
                                    payload.fecha_fin, payload.motivo, payload.horas)


# ---- Libro de horas (cuánto debe cada empleado, y cómo lo va pagando) ----

class NuevoMovimientoHorasRH(BaseModel):
    usuario_id: int
    tipo: str  # 'debe' o 'pago'
    horas: float
    notas: Optional[str] = None


@app.get("/api/rh/horas")
def api_listar_saldos_horas_rh(usuario: dict = Depends(requiere_admin_rh)):
    """Resumen de todos los empleados con movimientos — solo administrador."""
    return db.listar_saldos_horas_todos(usuario["empresa_id"])


@app.get("/api/rh/horas/pendientes-encargado")
def api_pagos_horas_pendientes_encargado(usuario: dict = Depends(requiere_encargado_sucursal)):
    # IMPORTANTE: esta ruta específica tiene que registrarse ANTES que
    # "/api/rh/horas/{usuario_id}" — si no, FastAPI intenta interpretar
    # "pendientes-encargado" como si fuera un usuario_id numérico y truena
    # con un error 422 antes de siquiera llegar aquí.
    mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
    if not mi_sucursal_id:
        return []
    return db.listar_pagos_horas_pendientes_encargado(usuario["empresa_id"], mi_sucursal_id)


@app.get("/api/rh/horas/{usuario_id}")
def api_consultar_horas_usuario(usuario_id: int, usuario: dict = Depends(requiere_empresa)):
    """Un empleado puede consultar SU PROPIO saldo; el administrador puede ver
    el de cualquiera; el encargado de sucursal puede ver el de su gente."""
    if usuario["rol"] != "admin" and usuario["id"] != usuario_id:
        puede_ver = False
        if usuario["rol"] == "encargado_sucursal":
            mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
            sucursal_de_la_persona = db.obtener_sucursal_id_usuario(usuario_id)
            puede_ver = bool(mi_sucursal_id) and sucursal_de_la_persona == mi_sucursal_id
        if not puede_ver:
            raise HTTPException(status_code=403, detail="No puedes consultar las horas de alguien más")
    saldo = db.saldo_horas_usuario(usuario["empresa_id"], usuario_id)
    movimientos = db.listar_movimientos_horas_rh(usuario["empresa_id"], usuario_id)
    return {**saldo, "movimientos": movimientos}


@app.get("/api/rh/horas-sucursal/saldos")
def api_saldos_horas_mi_sucursal(usuario: dict = Depends(requiere_empresa)):
    """Para el encargado de sucursal: quiénes de su gente deben horas ahorita
    (y quiénes ya van al corriente) — así sabe a quién darle seguimiento."""
    if usuario["rol"] != "encargado_sucursal":
        raise HTTPException(status_code=403, detail="Esto es solo para el encargado de sucursal")
    mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
    if not mi_sucursal_id:
        return []
    return db.listar_saldos_horas_sucursal(usuario["empresa_id"], mi_sucursal_id)


@app.post("/api/rh/horas/movimientos")
def api_registrar_movimiento_horas_rh(payload: NuevoMovimientoHorasRH, usuario: dict = Depends(requiere_admin_rh)):
    if payload.tipo not in db.TIPOS_MOVIMIENTO_HORAS_RH:
        raise HTTPException(status_code=400, detail="Tipo de movimiento inválido")
    if payload.horas <= 0:
        raise HTTPException(status_code=400, detail="Las horas deben ser un número positivo")
    objetivo = next((u for u in db.listar_usuarios(usuario["empresa_id"]) if u["id"] == payload.usuario_id), None)
    if not objetivo:
        raise HTTPException(status_code=404, detail="Usuario no encontrado en tu empresa")
    db.registrar_movimiento_horas_rh(usuario["empresa_id"], payload.usuario_id, payload.tipo, payload.horas,
                                      payload.notas, registrado_por_id=usuario["id"])
    saldo = db.saldo_horas_usuario(usuario["empresa_id"], payload.usuario_id)
    movimientos = db.listar_movimientos_horas_rh(usuario["empresa_id"], payload.usuario_id)
    return {**saldo, "movimientos": movimientos}


class SolicitudPagoHoras(BaseModel):
    fecha: str
    horas: float = Field(gt=0, le=24)
    motivo: str = Field(min_length=1, max_length=300)


@app.post("/api/rh/horas/solicitar-pago")
def api_solicitar_pago_horas(payload: SolicitudPagoHoras, usuario: dict = Depends(requiere_empresa)):
    """Cualquier persona puede registrar que 'pagó' horas (se quedó tiempo
    extra, trabajó parte de su comida, etc.) — queda pendiente de que el
    encargado de su sucursal lo autorice antes de que cuente en su saldo."""
    movimiento_id = db.solicitar_pago_horas_empleado(usuario["empresa_id"], usuario["id"], payload.fecha,
                                                       payload.horas, payload.motivo)
    return {"id": movimiento_id}


class FirmaAprobacionHoras(BaseModel):
    firma_base64: str = Field(min_length=100)


@app.post("/api/rh/horas/movimientos/{movimiento_id}/aprobar")
def api_aprobar_pago_horas(movimiento_id: int, payload: FirmaAprobacionHoras, usuario: dict = Depends(requiere_encargado_sucursal)):
    movimiento = db.obtener_movimiento_horas_rh(usuario["empresa_id"], movimiento_id)
    if not movimiento:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
    sucursal_de_la_persona = db.obtener_sucursal_id_usuario(movimiento["usuario_id"])
    if not mi_sucursal_id or sucursal_de_la_persona != mi_sucursal_id:
        raise HTTPException(status_code=403, detail="Esta persona no es de tu sucursal")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    if not db.aprobar_pago_horas(usuario["empresa_id"], movimiento_id, usuario["id"], payload.firma_base64):
        raise HTTPException(status_code=400, detail="Este pago ya no está pendiente de tu firma")
    return {"ok": True}


class RechazoPagoHoras(BaseModel):
    motivo: Optional[str] = None


@app.post("/api/rh/horas/movimientos/{movimiento_id}/rechazar")
def api_rechazar_pago_horas(movimiento_id: int, payload: RechazoPagoHoras, usuario: dict = Depends(requiere_encargado_sucursal)):
    movimiento = db.obtener_movimiento_horas_rh(usuario["empresa_id"], movimiento_id)
    if not movimiento:
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
    sucursal_de_la_persona = db.obtener_sucursal_id_usuario(movimiento["usuario_id"])
    if not mi_sucursal_id or sucursal_de_la_persona != mi_sucursal_id:
        raise HTTPException(status_code=403, detail="Esta persona no es de tu sucursal")
    if not db.rechazar_pago_horas(usuario["empresa_id"], movimiento_id, usuario["id"], payload.motivo):
        raise HTTPException(status_code=400, detail="Este pago ya no está pendiente de tu firma")
    return {"ok": True}


@app.delete("/api/rh/horas/movimientos/{movimiento_id}")
def api_eliminar_movimiento_horas_rh(movimiento_id: int, usuario: dict = Depends(requiere_admin_rh)):
    if not db.eliminar_movimiento_horas_rh(usuario["empresa_id"], movimiento_id):
        raise HTTPException(status_code=404, detail="Movimiento no encontrado")
    return {"ok": True}


# ---- Cursos por perfil de puesto ----

class NuevoCursoRH(BaseModel):
    nombre: str = Field(min_length=1, max_length=200)
    descripcion: Optional[str] = None
    puesto_objetivo: Optional[str] = None
    dias_duracion: Optional[int] = Field(default=None, ge=1, le=365)
    fecha_limite: Optional[str] = None


class ParticipanteCursoRH(BaseModel):
    usuario_id: int


class FirmaCursoRH(BaseModel):
    firma_base64: str = Field(min_length=100)
    evidencia_base64: str = Field(min_length=100)
    evidencia_nombre: Optional[str] = None


@app.get("/api/rh/cursos")
def api_listar_cursos_rh(usuario: dict = Depends(requiere_admin_rh)):
    return db.listar_cursos_rh(usuario["empresa_id"])


@app.post("/api/rh/cursos")
def api_crear_curso_rh(payload: NuevoCursoRH, usuario: dict = Depends(requiere_admin_rh)):
    curso_id = db.crear_curso_rh(usuario["empresa_id"], payload.nombre, payload.descripcion,
                                  payload.puesto_objetivo, payload.dias_duracion, payload.fecha_limite, usuario["id"])
    return db.obtener_curso_rh(usuario["empresa_id"], curso_id)


@app.get("/api/rh/cursos/{curso_id}")
def api_detalle_curso_rh(curso_id: int, usuario: dict = Depends(requiere_admin_rh)):
    curso = db.obtener_curso_rh(usuario["empresa_id"], curso_id)
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    return curso


@app.delete("/api/rh/cursos/{curso_id}")
def api_eliminar_curso_rh(curso_id: int, usuario: dict = Depends(requiere_admin_rh)):
    if not db.eliminar_curso_rh(usuario["empresa_id"], curso_id):
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    return {"ok": True}


@app.post("/api/rh/cursos/{curso_id}/participantes")
def api_agregar_participante_curso(curso_id: int, payload: ParticipanteCursoRH, usuario: dict = Depends(requiere_admin_rh)):
    curso = db.obtener_curso_rh(usuario["empresa_id"], curso_id)
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    objetivo = next((u for u in db.listar_usuarios(usuario["empresa_id"]) if u["id"] == payload.usuario_id), None)
    if not objetivo:
        raise HTTPException(status_code=404, detail="Usuario no encontrado en tu empresa")
    db.agregar_participante_curso(curso_id, payload.usuario_id, usuario["id"])
    return db.obtener_curso_rh(usuario["empresa_id"], curso_id)


@app.delete("/api/rh/cursos/{curso_id}/participantes/{usuario_id}")
def api_quitar_participante_curso(curso_id: int, usuario_id: int, usuario: dict = Depends(requiere_admin_rh)):
    if not db.obtener_curso_rh(usuario["empresa_id"], curso_id):
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    if not db.quitar_participante_curso(curso_id, usuario_id, usuario["id"]):
        raise HTTPException(status_code=400, detail="No se puede quitar: esta persona ya completó el curso (queda como historial)")
    return db.obtener_curso_rh(usuario["empresa_id"], curso_id)


@app.get("/api/rh/mis-cursos")
def api_mis_cursos_rh(usuario: dict = Depends(requiere_ver_rh)):
    return db.listar_cursos_usuario(usuario["empresa_id"], usuario["id"])


@app.post("/api/rh/cursos/{curso_id}/firmar")
def api_firmar_curso_rh(curso_id: int, payload: FirmaCursoRH, usuario: dict = Depends(requiere_ver_rh)):
    curso = db.obtener_curso_rh(usuario["empresa_id"], curso_id)
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    if not db.es_participante_curso(curso_id, usuario["id"]):
        raise HTTPException(status_code=403, detail="No estás asignado a este curso")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    if len(payload.evidencia_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La evidencia pesa demasiado")
    db.firmar_curso_rh(curso_id, usuario["id"], payload.firma_base64, payload.evidencia_base64, payload.evidencia_nombre)
    return {"ok": True}


@app.get("/api/rh/cursos/{curso_id}/constancia/{usuario_id}")
def api_constancia_curso_rh(curso_id: int, usuario_id: int, usuario: dict = Depends(requiere_ver_rh)):
    """El propio empleado descarga SU constancia; el administrador puede
    descargar la de cualquiera (por ejemplo, para reimprimirla)."""
    if usuario["rol"] != "admin" and usuario["id"] != usuario_id:
        raise HTTPException(status_code=403, detail="No puedes descargar la constancia de alguien más")
    curso = db.obtener_curso_rh(usuario["empresa_id"], curso_id)
    if not curso:
        raise HTTPException(status_code=404, detail="Curso no encontrado")
    participante = next((p for p in curso["participantes"] if p["usuario_id"] == usuario_id), None)
    if not participante or not participante.get("completado_en"):
        raise HTTPException(status_code=400, detail="Esta persona todavía no completó el curso")
    empresa = db.obtener_empresa(usuario["empresa_id"])
    pdf_bytes = pdfs_rh.generar_constancia_curso(curso, participante, empresa)
    nombre_archivo = f"constancia_{curso['nombre'][:30].replace(' ', '_')}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


NOMBRES_TIPO_INCIDENCIA_RH_PDF = {
    "dia_libre_sin_goce": "Día libre sin goce de sueldo", "enfermedad": "Falta por enfermedad",
    "lesion": "Lesión", "embarazo": "Embarazo", "accidente": "Accidente", "otro": "Otro",
}
NOMBRES_ESTADO_INCIDENCIA_RH_PDF = {"pendiente": "Pendiente", "aprobada": "Aprobada", "rechazada": "Rechazada"}


def _elementos_reporte_rh_por_empleado(empresa_id):
    """Arma las secciones del reporte de RH: una por empleado (orden alfabético),
    cada una con el detalle completo de sus incidencias — tipo, fechas, horas,
    motivo, estado y quién la resolvió."""
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle

    incidencias = db.listar_incidencias_rh(empresa_id)
    por_empleado = {}
    for i in incidencias:
        por_empleado.setdefault(i["usuario_nombre"], []).append(i)

    styles = getSampleStyleSheet()
    estilo_persona = ParagraphStyle("Persona", parent=styles["Heading3"], fontSize=12, textColor=colors.HexColor("#D8192F"),
                                     spaceBefore=14, spaceAfter=4)
    estilo_celda = ParagraphStyle("Celda", parent=styles["Normal"], fontSize=7.5, leading=9)

    elementos = [Paragraph(f"{len(incidencias)} incidencia(s) en total, de {len(por_empleado)} persona(s)", styles["Heading2"])]
    if not incidencias:
        elementos.append(Paragraph("No hay incidencias registradas.", styles["Normal"]))
        return elementos

    for nombre in sorted(por_empleado.keys()):
        lista = por_empleado[nombre]
        puesto = lista[0].get("usuario_puesto")
        elementos.append(Paragraph(f"{nombre}{f' — {puesto}' if puesto else ''} ({len(lista)})", estilo_persona))
        filas = []
        for i in lista:
            fechas = i["fecha_inicio"][:10]
            if i.get("fecha_fin") and i["fecha_fin"][:10] != fechas:
                fechas += f" al {i['fecha_fin'][:10]}"
            filas.append([
                Paragraph(NOMBRES_TIPO_INCIDENCIA_RH_PDF.get(i["tipo"], i["tipo"]), estilo_celda),
                fechas,
                f"{i['horas']} hrs" if i.get("horas") else "—",
                Paragraph((i.get("motivo") or "—")[:200], estilo_celda),
                NOMBRES_ESTADO_INCIDENCIA_RH_PDF.get(i["estado"], i["estado"]),
                i.get("resuelto_por_nombre") or "—",
            ])
        tabla = Table([["Tipo", "Fecha(s)", "Horas", "Motivo", "Estado", "Resuelto por"]] + filas,
                       colWidths=[3.3 * cm, 2.2 * cm, 1.3 * cm, 4.5 * cm, 2 * cm, 2.5 * cm])
        tabla.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#74767A")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 7.5), ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        elementos.append(tabla)
    return elementos


@app.get("/api/rh/reporte.pdf")
def api_reporte_rh_pdf(usuario: dict = Depends(requiere_admin_rh)):
    empresa = db.obtener_empresa(usuario["empresa_id"])
    elementos = _elementos_reporte_rh_por_empleado(usuario["empresa_id"])
    return _armar_pdf_simple("Reporte de Recursos Humanos", empresa["nombre"] if empresa else "", elementos, "incidencias_rh")


# ==================== REPARACIONES ====================

class NuevaSucursalReparacion(BaseModel):
    nombre: str = Field(min_length=1, max_length=120)
    prefijo: str = Field(min_length=1, max_length=10)
    departamento: Optional[str] = None
    telefonos: Optional[str] = None
    notas: Optional[str] = None


@app.get("/api/reparaciones/sucursales")
def api_listar_sucursales_reparacion(usuario: dict = Depends(requiere_empresa)):
    return db.listar_sucursales_reparacion(usuario["empresa_id"])


@app.post("/api/reparaciones/sucursales")
def api_crear_sucursal_reparacion(payload: NuevaSucursalReparacion, usuario: dict = Depends(requiere_admin_completo)):
    prefijo = re.sub(r"[^A-Za-z0-9]", "", payload.prefijo).upper()
    if not prefijo:
        raise HTTPException(status_code=400, detail="El prefijo debe tener al menos una letra o número")
    if payload.departamento:
        departamentos_validos = {d["nombre"] for d in db.listar_departamentos(usuario["empresa_id"])}
        if payload.departamento not in departamentos_validos:
            raise HTTPException(status_code=400, detail="Departamento inválido")
    sucursal_id = db.crear_sucursal_reparacion(usuario["empresa_id"], payload.nombre, prefijo, payload.departamento,
                                                payload.telefonos, payload.notas)
    return {"id": sucursal_id}


class ActualizacionSucursalReparacion(BaseModel):
    nombre: Optional[str] = None
    prefijo: Optional[str] = None
    departamento: Optional[str] = None
    activo: Optional[bool] = None
    telefonos: Optional[str] = None
    notas: Optional[str] = None


@app.patch("/api/reparaciones/sucursales/{sucursal_id}")
def api_actualizar_sucursal_reparacion(sucursal_id: int, payload: ActualizacionSucursalReparacion, usuario: dict = Depends(requiere_admin_completo)):
    if not db.obtener_sucursal_reparacion(usuario["empresa_id"], sucursal_id):
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    datos = payload.dict(exclude_unset=True)
    if "prefijo" in datos and datos["prefijo"]:
        prefijo = re.sub(r"[^A-Za-z0-9]", "", datos["prefijo"]).upper()
        if not prefijo:
            raise HTTPException(status_code=400, detail="El prefijo debe tener al menos una letra o número")
        datos["prefijo"] = prefijo
    if "departamento" in datos and datos["departamento"]:
        departamentos_validos = {d["nombre"] for d in db.listar_departamentos(usuario["empresa_id"])}
        if datos["departamento"] not in departamentos_validos:
            raise HTTPException(status_code=400, detail="Departamento inválido")
    return db.actualizar_sucursal_reparacion(usuario["empresa_id"], sucursal_id, **datos)


class NuevaReparacion(BaseModel):
    sucursal_id: int
    cliente_nombre: str = Field(min_length=1, max_length=160)
    cliente_telefono: Optional[str] = None
    equipo: Optional[str] = None
    marca: Optional[str] = None
    modelo: Optional[str] = None
    numero_serie: Optional[str] = None
    fecha_adquisicion: Optional[str] = None
    folio_adquisicion: Optional[str] = None
    garantia: bool = False
    falla_reportada: Optional[str] = None
    estado_fisico: Optional[str] = None
    accesorios_entregados: Optional[str] = None
    firma_recepcion: Optional[str] = None
    foto_estado_base64: Optional[str] = None
    foto_estado_nombre: Optional[str] = None
    departamento: str
    categoria: str


class ActualizacionReparacion(BaseModel):
    folio_microsip: Optional[str] = None
    cliente_nombre: Optional[str] = None
    cliente_telefono: Optional[str] = None
    asesor_recibe: Optional[str] = None
    equipo: Optional[str] = None
    marca: Optional[str] = None
    modelo: Optional[str] = None
    numero_serie: Optional[str] = None
    fecha_adquisicion: Optional[str] = None
    folio_adquisicion: Optional[str] = None
    garantia: Optional[bool] = None
    falla_reportada: Optional[str] = None
    estado_fisico: Optional[str] = None
    accesorios_entregados: Optional[str] = None
    diagnostico: Optional[str] = None
    autorizacion_precio: Optional[bool] = None
    autorizacion_medio: Optional[str] = None
    fecha_autorizacion: Optional[str] = None
    folio_solicitud_traspaso: Optional[str] = None
    costo_paqueteria: Optional[float] = None
    conclusion: Optional[str] = None
    recomendaciones: Optional[str] = None
    responsable_diagnostico_id: Optional[int] = None
    fecha_envio_proveedor: Optional[str] = None
    observaciones_entrega: Optional[str] = None
    firma_entrega: Optional[str] = None
    motivo_edicion: Optional[str] = None


class CambioEstadoReparacion(BaseModel):
    estado: str


class NuevoItemCosto(BaseModel):
    articulo: str = Field(min_length=1)
    cantidad: int = Field(default=1, ge=1)
    codigo: Optional[str] = None
    costo: float = Field(default=0, ge=0)


class NuevaEvidenciaReparacion(BaseModel):
    etapa: str = "ingreso"
    archivo_base64: str
    archivo_nombre: Optional[str] = None


class NuevaActualizacionReparacion(BaseModel):
    texto: str = Field(min_length=1)


@app.get("/api/reparaciones")
def api_listar_reparaciones(estado: Optional[str] = None, sucursal_id: Optional[int] = None, usuario: dict = Depends(requiere_ver_reparaciones)):
    creado_por_id = usuario["id"] if usuario["rol"] == "usuario" else None
    if usuario["rol"] == "almacen":
        # Un encargado de almacén solo ve reparaciones de SU propia sucursal, sin
        # importar qué sucursal_id le manden en la consulta (esto es seguridad, no solo filtro).
        sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
    return db.listar_reparaciones(usuario["empresa_id"], estado, sucursal_id, creado_por_id)


@app.post("/api/reparaciones")
def api_crear_reparacion(payload: NuevaReparacion, usuario: dict = Depends(requiere_empresa)):
    if not db.obtener_sucursal_reparacion(usuario["empresa_id"], payload.sucursal_id):
        raise HTTPException(status_code=404, detail="Sucursal no encontrada")
    departamentos_validos = {d["nombre"] for d in db.listar_departamentos(usuario["empresa_id"])}
    categorias_validas = {c["nombre"] for c in db.listar_categorias(usuario["empresa_id"])}
    if payload.departamento not in departamentos_validos:
        raise HTTPException(status_code=400, detail="Departamento inválido")
    if payload.categoria not in categorias_validas:
        raise HTTPException(status_code=400, detail="Categoría inválida")

    # Todos los campos de la orden de servicio son obligatorios (incluida la firma
    # del cliente y la foto del estado en que se recibe el equipo). El asesor NO se
    # pide como campo — siempre es quien tiene la sesión iniciada en este momento.
    campos_obligatorios = {
        "Teléfono del cliente": payload.cliente_telefono,
        "Tipo de equipo": payload.equipo, "Marca": payload.marca, "Modelo": payload.modelo,
        "Número de serie": payload.numero_serie, "Fecha de adquisición": payload.fecha_adquisicion,
        "Folio de adquisición": payload.folio_adquisicion,
        "Falla reportada": payload.falla_reportada, "Estado físico": payload.estado_fisico,
        "Accesorios entregados": payload.accesorios_entregados,
    }
    faltantes = [nombre for nombre, valor in campos_obligatorios.items() if not (valor and valor.strip())]
    if faltantes:
        raise HTTPException(status_code=400, detail=f"Faltan campos obligatorios: {', '.join(faltantes)}")
    if not re.match(r"^\d{10}$", payload.cliente_telefono.strip()):
        raise HTTPException(status_code=400, detail="El teléfono debe ser un número de exactamente 10 dígitos")
    if not payload.firma_recepcion:
        raise HTTPException(status_code=400, detail="Falta la firma del cliente")
    if not payload.foto_estado_base64:
        raise HTTPException(status_code=400, detail="Falta la foto del estado en que se recibe el equipo")
    if payload.firma_recepcion and len(payload.firma_recepcion) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    if payload.foto_estado_base64 and len(payload.foto_estado_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La foto pesa demasiado (máximo 5MB)")

    reparacion = db.crear_reparacion(
        usuario["empresa_id"], payload.sucursal_id, payload.cliente_nombre, payload.cliente_telefono.strip(),
        usuario["nombre"], payload.equipo, payload.marca, payload.modelo, payload.numero_serie,
        payload.fecha_adquisicion, payload.folio_adquisicion, payload.garantia, payload.falla_reportada,
        payload.estado_fisico, payload.accesorios_entregados, payload.firma_recepcion, payload.departamento,
        payload.categoria, usuario["id"], payload.foto_estado_base64, payload.foto_estado_nombre,
    )
    db.agregar_actualizacion_reparacion(
        reparacion["id"], usuario["id"],
        f"Se creó la orden de servicio — el cliente ({payload.cliente_nombre.strip()}) firmó de recibido.",
    )
    tecnicos = db.listar_tecnicos_activos(usuario["empresa_id"])
    ticket = db.obtener_ticket(reparacion["ticket_id"])
    notifications.notificar_nuevo_ticket(tecnicos, ticket)
    return reparacion


@app.get("/api/reparaciones/reporte.pdf")
def reporte_reparaciones_pdf(usuario: dict = Depends(requiere_staff)):
    from io import BytesIO
    from datetime import datetime as dt

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    empresa = db.obtener_empresa(usuario["empresa_id"])
    reparaciones = db.listar_reparaciones(usuario["empresa_id"])

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Reparaciones — {empresa['nombre'] if empresa else ''}", styles["Title"]),
        Paragraph(f"Generado el {dt.now().strftime('%d/%m/%Y %H:%M')} — {len(reparaciones)} reparación(es)", styles["Normal"]),
        Spacer(1, 16),
    ]

    datos = [["Folio", "Sucursal", "Cliente", "Equipo", "Estado", "Técnico", "Costo total"]]
    for r in reparaciones:
        datos.append([
            r["folio"], r.get("sucursal_nombre") or "—", r["cliente_nombre"], r.get("equipo") or "—",
            NOMBRES_ESTADO_REPARACION_PDF.get(r["estado"], r["estado"]),
            r.get("tecnico_nombre") or "sin asignar", f"${r.get('costo_total', 0):,.2f}",
        ])
    tabla = Table(datos, repeatRows=1)
    tabla.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D8192F")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
    ]))
    elementos.append(tabla)
    if not reparaciones:
        elementos.append(Paragraph("No hay reparaciones registradas.", styles["Normal"]))

    doc.build(elementos)
    buffer.seek(0)
    nombre_archivo = f"reparaciones_{dt.now().strftime('%Y%m%d')}.pdf"
    return Response(content=buffer.read(), media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"})


@app.get("/api/reparaciones/reporte.xlsx")
def reporte_reparaciones_xlsx(usuario: dict = Depends(requiere_staff)):
    from io import BytesIO
    from datetime import datetime as dt

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    reparaciones = db.listar_reparaciones(usuario["empresa_id"])

    wb = Workbook()
    ws = wb.active
    ws.title = "Reparaciones"
    encabezados = ["Folio", "Folio Microsip", "Sucursal", "Cliente", "Teléfono", "Equipo", "Marca", "Modelo",
                   "Estado", "Técnico", "Días transcurridos", "Costo total", "Fecha recepción", "Fecha entrega"]
    ws.append(encabezados)
    for col_idx, _ in enumerate(encabezados, start=1):
        celda = ws.cell(row=1, column=col_idx)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill(start_color="D8192F", end_color="D8192F", fill_type="solid")
        celda.alignment = Alignment(horizontal="center")

    for r in reparaciones:
        ws.append([
            r["folio"], r.get("folio_microsip") or "", r.get("sucursal_nombre") or "", r["cliente_nombre"],
            r.get("cliente_telefono") or "", r.get("equipo") or "", r.get("marca") or "", r.get("modelo") or "",
            NOMBRES_ESTADO_REPARACION_PDF.get(r["estado"], r["estado"]), r.get("tecnico_nombre") or "",
            r.get("dias_transcurridos") if r.get("dias_transcurridos") is not None else "",
            r.get("costo_total", 0), (r.get("fecha_recepcion") or "")[:10], (r.get("fecha_entrega") or "")[:10],
        ])

    for col_idx, encabezado in enumerate(encabezados, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(encabezado), 14) + 4
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    nombre_archivo = f"reparaciones_{dt.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buffer.read(), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"},
    )


@app.get("/api/reparaciones/{reparacion_id}")
def api_detalle_reparacion(reparacion_id: int, usuario: dict = Depends(requiere_ver_reparaciones)):
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "usuario" and reparacion["creado_por_id"] != usuario["id"]:
        raise HTTPException(status_code=403, detail="No puedes ver esta reparación")
    if usuario["rol"] == "almacen" and reparacion["sucursal_id"] != db.obtener_sucursal_id_usuario(usuario["id"]):
        raise HTTPException(status_code=403, detail="Esta reparación no es de tu sucursal")
    return reparacion


@app.patch("/api/reparaciones/{reparacion_id}")
def api_actualizar_reparacion(reparacion_id: int, payload: ActualizacionReparacion, usuario: dict = Depends(requiere_staff)):
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    enviados = payload.dict(exclude_unset=True)
    print(f"[DEBUG PATCH reparacion #{reparacion_id}] usuario={usuario.get('nombre')} (rol={usuario['rol']}) enviados={enviados}")
    motivo_edicion = (enviados.pop("motivo_edicion", None) or "").strip()
    toca_campos_tecnico = any(k in db._CAMPOS_TECNICO_REPARACION for k in enviados)
    # El bloqueo se activa específicamente cuando se guarda TEXTO real de
    # diagnóstico — no con cualquier otro campo de la misma pestaña (autorización,
    # folio de traspaso, costos) — para no bloquearlo antes de que el técnico
    # siquiera alcance a escribir el diagnóstico la primera vez.
    se_guarda_diagnostico_real = bool(enviados.get("diagnostico"))

    # No se puede empezar a diagnosticar mientras la reparación siga en "nueva"
    # (recién creada, todavía sin confirmar que se recibió) — aplica para
    # cualquier rol, incluido el administrador, es un orden de flujo, no un permiso.
    if toca_campos_tecnico and reparacion["estado"] == "nueva":
        raise HTTPException(status_code=400, detail="Primero cambia el estado a 'Recibido en diagnóstico' antes de llenar el diagnóstico")

    if usuario["rol"] == "tecnico":
        if reparacion.get("firma_salida_en"):
            raise HTTPException(status_code=403, detail="Esta reparación ya salió del taller — ya no puedes modificarla")

        # El técnico no puede tocar lo que la sucursal capturó al recibir el equipo
        # (cliente, equipo, falla reportada, accesorios, etc.) — solo su propio trabajo.
        campos_no_permitidos = [k for k in enviados if k in db._CAMPOS_RECEPCION_REPARACION]
        if campos_no_permitidos:
            raise HTTPException(status_code=403, detail="No puedes editar los datos de recepción capturados por la sucursal")

        # Una vez que el técnico guarda el diagnóstico, queda bloqueado — para
        # volver a tocar CUALQUIER campo de esta sección tiene que dar un motivo,
        # y eso se anota en la bitácora. Cambiar el ESTADO nunca pasa por aquí
        # (usa un endpoint aparte), así que el técnico siempre puede seguir
        # moviendo el estado sin ninguna restricción.
        if toca_campos_tecnico and reparacion.get("diagnostico_bloqueado"):
            if not motivo_edicion:
                raise HTTPException(status_code=400, detail="El diagnóstico ya está guardado — indica el motivo del cambio para poder editarlo")
            db.agregar_actualizacion_reparacion(
                reparacion_id, usuario["id"],
                f"Editó el diagnóstico (ya estaba guardado). Motivo: {motivo_edicion}",
            )

        db.actualizar_reparacion(usuario["empresa_id"], reparacion_id, campos_permitidos=db._CAMPOS_TECNICO_REPARACION, **enviados)
        _rep_tras_guardar = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
        print(f"[DEBUG PATCH reparacion #{reparacion_id}] tras guardar -> diagnostico={_rep_tras_guardar.get('diagnostico')!r} "
              f"autorizacion_precio={_rep_tras_guardar.get('autorizacion_precio')!r} autorizacion_medio={_rep_tras_guardar.get('autorizacion_medio')!r}")

        if se_guarda_diagnostico_real and not reparacion.get("diagnostico_bloqueado"):
            db.actualizar_reparacion(usuario["empresa_id"], reparacion_id,
                                      campos_permitidos=["diagnostico_bloqueado"], diagnostico_bloqueado=True)
            db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"],
                                                 "Guardó el diagnóstico — queda bloqueado para futuras ediciones.")
    else:
        # El administrador siempre puede editar sin necesidad de motivo, pero si el
        # diagnóstico ya estaba bloqueado, igual queda anotado en la bitácora.
        if toca_campos_tecnico and reparacion.get("diagnostico_bloqueado"):
            nota = "El administrador editó el diagnóstico (ya estaba guardado por el técnico)."
            if motivo_edicion:
                nota += f" Motivo: {motivo_edicion}"
            db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"], nota)
        db.actualizar_reparacion(usuario["empresa_id"], reparacion_id, **enviados)
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


@app.delete("/api/reparaciones/{reparacion_id}")
def api_eliminar_reparacion(reparacion_id: int, usuario: dict = Depends(requiere_admin_completo)):
    if not db.eliminar_reparacion(usuario["empresa_id"], reparacion_id):
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    return {"ok": True}


@app.patch("/api/reparaciones/{reparacion_id}/estado")
def api_cambiar_estado_reparacion(reparacion_id: int, payload: CambioEstadoReparacion, usuario: dict = Depends(requiere_staff)):
    if payload.estado not in db.ESTADOS_REPARACION:
        raise HTTPException(status_code=400, detail="Estado inválido")
    if payload.estado in ("envio_sucursal", "en_traslado", "listo_entrega"):
        raise HTTPException(status_code=400, detail="Este paso requiere una firma — usa 'Firmar salida', 'Firmar entrega al chofer' o 'Firmar ingreso a sucursal'")
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "tecnico" and reparacion.get("firma_salida_en"):
        raise HTTPException(status_code=403, detail="Esta reparación ya salió del taller — ya no puedes modificarla")
    db.cambiar_estado_reparacion(usuario["empresa_id"], reparacion_id, payload.estado)
    nombre_estado = NOMBRES_ESTADO_REPARACION_BITACORA.get(payload.estado, payload.estado)
    db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"], f"Cambió el estado a: {nombre_estado}")
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


class FirmaSalidaReparacion(BaseModel):
    firma_base64: str = Field(min_length=100)


@app.post("/api/reparaciones/{reparacion_id}/firma-salida")
def api_firmar_salida_reparacion(reparacion_id: int, payload: FirmaSalidaReparacion, usuario: dict = Depends(requiere_staff)):
    """El técnico (o admin) firma que el equipo sale del taller rumbo a la sucursal.
    A partir de este momento, el técnico ya no puede modificar NADA de la reparación
    — ni el diagnóstico, ni el estado — solo el administrador."""
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if reparacion["estado"] in ("envio_sucursal", "en_traslado", "listo_entrega", "entregado", "cancelado"):
        raise HTTPException(status_code=400, detail="Esta reparación ya pasó por este paso")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    db.firmar_salida_reparacion(usuario["empresa_id"], reparacion_id, usuario["id"], payload.firma_base64)
    db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"], "Firmó la salida del taller rumbo a la sucursal.")
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


class FirmaChoferReparacion(BaseModel):
    chofer_nombre: str = Field(min_length=1, max_length=160)
    firma_base64: str = Field(min_length=100)


@app.post("/api/reparaciones/{reparacion_id}/firma-chofer")
def api_firmar_chofer_reparacion(reparacion_id: int, payload: FirmaChoferReparacion, usuario: dict = Depends(requiere_staff)):
    """El chofer que se lleva el equipo firma de recibido — avanza el estado a
    'en_traslado'. Solo aplica justo después de la firma de salida.
    Exclusivo del administrador: el técnico ya queda bloqueado en cuanto firma la
    salida (justo lo que hace posible este paso), así que nunca llega a hacerlo él."""
    if usuario["rol"] != "admin":
        raise HTTPException(status_code=403, detail="Solo el administrador puede registrar la entrega al chofer")
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if reparacion["estado"] != "envio_sucursal":
        raise HTTPException(status_code=400, detail="Todavía falta la firma de salida, o el chofer ya firmó")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    db.firmar_chofer_reparacion(usuario["empresa_id"], reparacion_id, usuario["id"], payload.chofer_nombre.strip(), payload.firma_base64)
    db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"],
                                         f"Se entregó al chofer {payload.chofer_nombre.strip()} para su traslado a la sucursal.")
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


class FirmaIngresoReparacion(BaseModel):
    firma_base64: str = Field(min_length=100)


@app.post("/api/reparaciones/{reparacion_id}/firma-ingreso")
def api_firmar_ingreso_reparacion(reparacion_id: int, payload: FirmaIngresoReparacion, usuario: dict = Depends(requiere_ver_reparaciones)):
    """Recepción en la sucursal — el encargado de almacén de esa misma sucursal
    (identificada por el folio) siempre puede hacerlo. El administrador TAMBIÉN
    puede recibir cualquier reparación, sin importar la sucursal — por si hace
    falta cubrir cuando no hay alguien de almacén disponible."""
    if usuario["rol"] not in ("almacen", "admin"):
        raise HTTPException(status_code=403, detail="Solo un encargado de almacén o un administrador puede firmar la recepción")
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "almacen":
        mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
        if not mi_sucursal_id or reparacion["sucursal_id"] != mi_sucursal_id:
            raise HTTPException(status_code=403, detail="Esta reparación no es de tu sucursal — no puedes recibirla")
    if reparacion["estado"] != "en_traslado":
        raise HTTPException(status_code=400, detail="Esta reparación todavía no va en camino (falta la firma del chofer), o ya fue recibida")
    if len(payload.firma_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")
    db.firmar_ingreso_reparacion(usuario["empresa_id"], reparacion_id, usuario["id"], payload.firma_base64)
    db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"], "Confirmó la recepción del equipo en la sucursal.")
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


class EntregaReparacion(BaseModel):
    observaciones_entrega: Optional[str] = None
    firma_entrega: Optional[str] = None


@app.post("/api/reparaciones/{reparacion_id}/entregar")
def api_entregar_reparacion(reparacion_id: int, payload: EntregaReparacion, usuario: dict = Depends(requiere_ver_reparaciones)):
    """Registra la entrega al cliente y cierra la reparación. El staff puede usarlo
    siempre; un empleado solo puede entregar SU PROPIA reparación, y solo cuando ya
    está en 'Listo para entrega' (el almacén ya la recibió en su sucursal). El
    encargado de almacén SOLO hace esto (entregar) — nada más del proceso — y
    únicamente para reparaciones de su propia sucursal."""
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "usuario":
        if reparacion["creado_por_id"] != usuario["id"]:
            raise HTTPException(status_code=403, detail="No puedes ver esta reparación")
        if reparacion["estado"] != "listo_entrega":
            raise HTTPException(status_code=400, detail="Esta reparación todavía no está lista para entregar (falta que el almacén la reciba)")
    if usuario["rol"] == "almacen":
        mi_sucursal_id = db.obtener_sucursal_id_usuario(usuario["id"])
        if not mi_sucursal_id or reparacion["sucursal_id"] != mi_sucursal_id:
            raise HTTPException(status_code=403, detail="Esta reparación no es de tu sucursal — no puedes entregarla")
        if reparacion["estado"] != "listo_entrega":
            raise HTTPException(status_code=400, detail="Esta reparación todavía no está lista para entregar")
    if payload.firma_entrega and len(payload.firma_entrega) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="La firma pesa demasiado")

    campos = {}
    if payload.observaciones_entrega is not None:
        campos["observaciones_entrega"] = payload.observaciones_entrega
    if payload.firma_entrega is not None:
        campos["firma_entrega"] = payload.firma_entrega
    if campos:
        db.actualizar_reparacion(usuario["empresa_id"], reparacion_id, **campos)
    db.cambiar_estado_reparacion(usuario["empresa_id"], reparacion_id, "entregado")
    db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"], "Registró la entrega del equipo al cliente.")
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


@app.post("/api/reparaciones/{reparacion_id}/items-costo")
def api_agregar_item_costo(reparacion_id: int, payload: NuevoItemCosto, usuario: dict = Depends(requiere_staff)):
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "tecnico" and reparacion.get("firma_salida_en"):
        raise HTTPException(status_code=403, detail="Esta reparación ya salió del taller — ya no puedes modificarla")
    db.agregar_item_costo(reparacion_id, payload.articulo, payload.cantidad, payload.codigo, payload.costo)
    db.agregar_actualizacion_reparacion(
        reparacion_id, usuario["id"],
        f"Agregó al costo: {payload.articulo} — {payload.cantidad} x ${payload.costo:,.2f} = ${payload.cantidad * payload.costo:,.2f}",
    )
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


@app.delete("/api/reparaciones/items-costo/{item_id}")
def api_eliminar_item_costo(item_id: int, usuario: dict = Depends(requiere_staff)):
    reparacion_id = db.obtener_reparacion_id_de_item_costo(item_id)
    if reparacion_id:
        reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
        if reparacion and usuario["rol"] == "tecnico" and reparacion.get("firma_salida_en"):
            raise HTTPException(status_code=403, detail="Esta reparación ya salió del taller — ya no puedes modificarla")
    db.eliminar_item_costo(item_id)
    if reparacion_id:
        db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"], "Quitó un ítem del costo de la reparación.")
    return {"ok": True}


@app.post("/api/reparaciones/{reparacion_id}/evidencias")
def api_agregar_evidencia_reparacion(reparacion_id: int, payload: NuevaEvidenciaReparacion, usuario: dict = Depends(requiere_staff)):
    if not db.obtener_reparacion(usuario["empresa_id"], reparacion_id):
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if len(payload.archivo_base64) > MAX_ADJUNTO_BASE64:
        raise HTTPException(status_code=400, detail="El archivo pesa demasiado (máximo 5MB)")
    db.agregar_evidencia_reparacion(reparacion_id, payload.etapa, payload.archivo_base64, payload.archivo_nombre, usuario["id"])
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


@app.post("/api/reparaciones/{reparacion_id}/actualizaciones")
def api_agregar_actualizacion_reparacion(reparacion_id: int, payload: NuevaActualizacionReparacion, usuario: dict = Depends(requiere_staff)):
    if not db.obtener_reparacion(usuario["empresa_id"], reparacion_id):
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    db.agregar_actualizacion_reparacion(reparacion_id, usuario["id"], payload.texto)
    return db.obtener_reparacion(usuario["empresa_id"], reparacion_id)


@app.get("/api/reparaciones/{reparacion_id}/orden-servicio.pdf")
def api_pdf_orden_servicio(reparacion_id: int, usuario: dict = Depends(requiere_empresa)):
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "usuario" and reparacion["creado_por_id"] != usuario["id"]:
        raise HTTPException(status_code=403, detail="No puedes ver esta reparación")
    empresa = db.obtener_empresa(usuario["empresa_id"])
    pdf_bytes = pdfs_reparaciones.generar_orden_servicio(reparacion, empresa)
    return Response(content=pdf_bytes, media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename=orden_servicio_{reparacion['folio']}.pdf"})


@app.get("/api/reparaciones/{reparacion_id}/diagnostico.pdf")
def api_pdf_diagnostico(reparacion_id: int, usuario: dict = Depends(requiere_empresa)):
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "usuario" and reparacion["creado_por_id"] != usuario["id"]:
        raise HTTPException(status_code=403, detail="No puedes ver esta reparación")
    empresa = db.obtener_empresa(usuario["empresa_id"])
    pdf_bytes = pdfs_reparaciones.generar_diagnostico(reparacion, empresa)
    return Response(content=pdf_bytes, media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename=diagnostico_{reparacion['folio']}.pdf"})


@app.get("/api/reparaciones/{reparacion_id}/conformidad-entrega.pdf")
def api_pdf_conformidad_entrega(reparacion_id: int, usuario: dict = Depends(requiere_empresa)):
    reparacion = db.obtener_reparacion(usuario["empresa_id"], reparacion_id)
    if not reparacion:
        raise HTTPException(status_code=404, detail="Reparación no encontrada")
    if usuario["rol"] == "usuario" and reparacion["creado_por_id"] != usuario["id"]:
        raise HTTPException(status_code=403, detail="No puedes ver esta reparación")
    empresa = db.obtener_empresa(usuario["empresa_id"])
    pdf_bytes = pdfs_reparaciones.generar_conformidad_entrega(reparacion, empresa)
    return Response(content=pdf_bytes, media_type="application/pdf",
                     headers={"Content-Disposition": f"attachment; filename=conformidad_entrega_{reparacion['folio']}.pdf"})


NOMBRES_ESTADO_REPARACION_PDF = {
    "en_diagnostico": "En diagnóstico", "esperando_autorizacion": "Esperando autorización",
    "en_reparacion": "En reparación", "con_proveedor": "Con proveedor", "esperando_refaccion": "Esperando refacción",
    "control_calidad": "Control de calidad", "envio_sucursal": "Envío a sucursal", "listo_entrega": "Listo para entrega",
    "entregado": "Entregado", "cancelado": "Cancelado",
}


# ==================== BORRADO MASIVO ====================

@app.get("/api/admin/borrado-masivo/contar")
def api_contar_borrado_masivo(tabla: str, fecha_desde: str, fecha_hasta: str, usuario: dict = Depends(requiere_admin_completo)):
    if tabla not in db.TABLAS_BORRADO_MASIVO:
        raise HTTPException(status_code=400, detail="Tabla inválida")
    cantidad = db.contar_registros_borrado_masivo(usuario["empresa_id"], tabla, fecha_desde, fecha_hasta)
    return {"cantidad": cantidad}


class BorradoMasivo(BaseModel):
    tabla: str
    fecha_desde: str
    fecha_hasta: str
    confirmacion: str


@app.post("/api/admin/borrado-masivo")
def api_borrado_masivo(payload: BorradoMasivo, usuario: dict = Depends(requiere_admin_completo)):
    if payload.tabla not in db.TABLAS_BORRADO_MASIVO:
        raise HTTPException(status_code=400, detail="Tabla inválida")
    if payload.confirmacion.strip().upper() != "BORRAR":
        raise HTTPException(status_code=400, detail="Debes escribir BORRAR para confirmar")
    eliminados = db.borrar_masivo(usuario["empresa_id"], payload.tabla, payload.fecha_desde, payload.fecha_hasta)
    return {"eliminados": eliminados}


# ==================== MICROSIP (conexión de solo lectura a Firebird) ====================

def _requiere_microsip_disponible():
    if not MICROSIP_DISPONIBLE:
        raise HTTPException(
            status_code=503,
            detail="El driver de Firebird ('fdb') todavía no está instalado en el servidor — agrégalo a requirements.txt y redespliega.",
        )


class ConfigMicrosip(BaseModel):
    host: str = Field(min_length=1)
    puerto: int = Field(default=3050, ge=1, le=65535)
    ruta_db: str = Field(min_length=1)
    usuario: str = Field(min_length=1)
    password: Optional[str] = None  # None = no cambiarla; "" = borrarla


@app.get("/api/microsip/config")
def api_obtener_config_microsip(usuario: dict = Depends(requiere_admin_completo)):
    config = db.obtener_config_microsip_publica(usuario["empresa_id"])
    return config or {"microsip_host": None, "microsip_puerto": 3050, "microsip_ruta_db": None,
                       "microsip_usuario": None, "tiene_password": False}


@app.post("/api/microsip/config")
def api_guardar_config_microsip(payload: ConfigMicrosip, usuario: dict = Depends(requiere_admin_completo)):
    db.actualizar_config_microsip(usuario["empresa_id"], payload.host, payload.puerto, payload.ruta_db,
                                   payload.usuario, payload.password)
    return {"ok": True}


@app.post("/api/microsip/probar-conexion")
def api_probar_conexion_microsip(usuario: dict = Depends(requiere_admin_completo)):
    _requiere_microsip_disponible()
    config = db.obtener_config_microsip(usuario["empresa_id"])
    ok, mensaje = microsip.probar_conexion(config)
    if not ok:
        raise HTTPException(status_code=400, detail=mensaje)
    return {"ok": True, "mensaje": mensaje}


@app.get("/api/microsip/tablas")
def api_listar_tablas_microsip(usuario: dict = Depends(requiere_admin_completo)):
    _requiere_microsip_disponible()
    config = db.obtener_config_microsip(usuario["empresa_id"])
    try:
        return {"tablas": microsip.listar_tablas(config)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/microsip/tablas/{tabla}/columnas")
def api_listar_columnas_microsip(tabla: str, usuario: dict = Depends(requiere_admin_completo)):
    _requiere_microsip_disponible()
    config = db.obtener_config_microsip(usuario["empresa_id"])
    try:
        return {"columnas": microsip.listar_columnas(config, tabla)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/microsip/tablas/{tabla}/muestra")
def api_muestra_tabla_microsip(tabla: str, limite: int = 20, usuario: dict = Depends(requiere_admin_completo)):
    _requiere_microsip_disponible()
    config = db.obtener_config_microsip(usuario["empresa_id"])
    try:
        return {"filas": microsip.consultar_muestra(config, tabla, limite)}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# ==================== FRONTEND ESTÁTICO ====================

FRONTEND_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend")
app.mount("/static", StaticFiles(directory=FRONTEND_DIR), name="static")


@app.get("/")
def index():
    return FileResponse(os.path.join(FRONTEND_DIR, "index.html"))


@app.get("/sw.js")
def service_worker():
    return FileResponse(os.path.join(FRONTEND_DIR, "sw.js"), media_type="application/javascript")
