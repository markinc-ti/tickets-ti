"""
Importar reparaciones históricas desde el Excel que usaban antes de tener
la app (hoja 'Area de Reparacione' del formato de Mark·Inc).

Flujo de dos pasos, pensado para que el administrador revise antes de
guardar nada:
  1. previsualizar(): lee el Excel, arma cada fila candidata, intenta
     encontrar la sucursal por el prefijo del folio o por el texto de la
     columna Sucursal, y marca las que ya existen (mismo folio) como
     duplicadas. No toca la base de datos.
  2. Una vez que el frontend le confirma sucursal_id para las filas que
     quedaron sin match y qué filas sí quiere importar, se llama a
     db.importar_reparaciones_lote() con el resultado ya limpio.
"""
import io
import re
import openpyxl

# Encabezados esperados en la hoja 'Area de Reparacione' — si el archivo
# viene con nombres ligeramente distintos, ESTADOS_MAP y las columnas de
# abajo son lo primero que hay que ajustar.
ESTADOS_MAP = {
    "entregado": "entregado",
    "cancelado": "cancelado",
    "con proveedor": "con_proveedor",
    "esperando autorización": "esperando_autorizacion",
    "esperando autorizacion": "esperando_autorizacion",
    "en diagnóstico": "en_diagnostico",
    "en diagnostico": "en_diagnostico",
    "en reparación": "en_reparacion",
    "en reparacion": "en_reparacion",
    "esperando refacción": "esperando_refaccion",
    "esperando refaccion": "esperando_refaccion",
    "control de calidad": "control_calidad",
    "listo para entrega": "listo_entrega",
}


def _texto(valor):
    if valor is None:
        return None
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    texto = str(valor).strip()
    return texto or None


def _fecha_iso(valor):
    """Regresa una fecha en formato ISO (o None) — nunca un texto crudo sin
    validar, porque eso después hace tronar datetime.fromisoformat() en el
    resto de la app (esto fue justo lo que rompió el listado completo de
    reparaciones la primera vez: una celda de fecha capturada como texto
    libre, tipo '26/01/2026', se guardó tal cual y crasheaba la consulta)."""
    if valor is None or valor == "":
        return None
    if hasattr(valor, "isoformat"):
        return valor.isoformat()
    texto = _texto(valor)
    if not texto:
        return None
    try:
        from datetime import datetime as _dt
        return _dt.fromisoformat(texto).isoformat()
    except ValueError:
        pass
    # Intenta los formatos de fecha más comunes en un Excel mexicano
    # (día/mes/año, con o sin hora).
    for formato in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y"):
        try:
            from datetime import datetime as _dt
            return _dt.strptime(texto, formato).isoformat()
        except ValueError:
            continue
    # No se pudo interpretar como fecha — mejor guardar nada que guardar
    # basura que después rompa el resto de la app.
    return None


def _booleano_si_no(valor):
    texto = _texto(valor)
    if not texto:
        return None
    return texto.strip().lower() in ("si", "sí", "s", "true", "1")


def _prefijo_folio(folio):
    m = re.match(r"^([A-Za-z]+)", (folio or "").strip())
    return m.group(1).upper() if m else None


def _encontrar_sucursal(prefijo_folio, texto_sucursal, sucursales):
    """sucursales: lista de dicts {id, nombre, prefijo}. Intenta por
    prefijo de folio primero (más confiable), luego por nombre."""
    if prefijo_folio:
        for s in sucursales:
            if s["prefijo"].upper() == prefijo_folio:
                return s["id"]
    if texto_sucursal:
        texto_norm = texto_sucursal.strip().lower()
        for s in sucursales:
            if s["nombre"].strip().lower() == texto_norm:
                return s["id"]
        # Coincidencia parcial como último intento (ej. "corpo" vs "Corporativo")
        for s in sucursales:
            if texto_norm in s["nombre"].strip().lower() or s["nombre"].strip().lower() in texto_norm:
                return s["id"]
    return None


def previsualizar(archivo_bytes: bytes, sucursales: list, folios_existentes: set):
    """Lee el Excel y regresa una lista de filas candidatas para el
    frontend, cada una con un 'estatus': 'lista', 'duplicada' o
    'sin_sucursal' — no escribe nada en la base de datos todavía."""
    wb = openpyxl.load_workbook(io.BytesIO(archivo_bytes), data_only=True)
    if "Area de Reparacione" not in wb.sheetnames:
        raise ValueError(
            "No encontré la hoja 'Area de Reparacione' en este archivo. "
            f"Hojas encontradas: {', '.join(wb.sheetnames)}"
        )
    ws = wb["Area de Reparacione"]

    encabezados = [(_texto(c.value) or "") for c in ws[1]]

    def col(nombre):
        for i, h in enumerate(encabezados):
            if h.strip().lower() == nombre.strip().lower():
                return i
        return None

    idx = {
        "estado": col("Estado"), "fecha": col("Fecha"), "folio": col("Folio"),
        "sucursal": col("Sucursal"), "cliente": col("Cliente"), "telefono": col("Teléfono"),
        "equipo": col("Equipo"), "marca": col("Marca"), "modelo": col("Modelo"), "ns": col("NS"),
        "garantia": col("Garantia "), "problema": col("PROBLEMA PRESENTADO"),
        "fecha_recepcion": col("Fecha de Recepción "), "autorizacion": col("Autorización Precio por el cliente"),
        "diagnostico": col("Diagnostico (Qué se le hizo al equipo)"),
        "folio_traspaso": col("Folio de Solicitud y Traspaso"), "costo_refaccion": col("Costo Refacción "),
        "servicio": col("Servicio"), "paqueteria": col("Paquetería "), "fecha_salida": col("Fecha de Salida"),
    }
    if idx["folio"] is None or idx["cliente"] is None:
        raise ValueError("El Excel no tiene las columnas esperadas (Folio, Cliente, ...) — revisa que sea el formato correcto.")

    filas_procesadas = []
    folios_vistos_en_este_archivo = {}  # folio_original -> lista de fingerprints ya vistos

    def _fingerprint(fila_datos):
        return (
            (fila_datos.get("falla_reportada") or "").strip().lower(),
            (fila_datos.get("diagnostico") or "").strip().lower(),
            round(sum(i["costo"] for i in fila_datos["items_costo"]) + fila_datos.get("costo_paqueteria", 0), 2),
        )

    for fila in ws.iter_rows(min_row=2, values_only=True):
        folio = _texto(fila[idx["folio"]]) if idx["folio"] is not None else None
        cliente = _texto(fila[idx["cliente"]]) if idx["cliente"] is not None else None
        sucursal_texto = _texto(fila[idx["sucursal"]]) if idx["sucursal"] is not None else None
        if not folio or not cliente:
            continue  # fila vacía o encabezado repetido

        # Algunos de estos formatos traen la fila de encabezado repetida
        # como si fuera un dato más (ej. Folio = "Folio" literal).
        if folio.strip().lower() == "folio":
            continue

        # Caso puntual visto en el archivo original: en algunas filas el
        # nombre del cliente y el de la sucursal quedaron capturados al
        # revés (columna Sucursal = nombre de persona, columna Cliente =
        # nombre de sucursal). Si el "cliente" es en realidad el nombre de
        # una sucursal conocida y la "sucursal" no matchea con ninguna,
        # los intercambiamos.
        if sucursal_texto and cliente and _encontrar_sucursal(None, cliente, sucursales) and not _encontrar_sucursal(_prefijo_folio(folio), sucursal_texto, sucursales):
            cliente, sucursal_texto = sucursal_texto, cliente

        # Fila con folio/sucursal/cliente idénticos → dato basura de captura.
        if folio.strip().lower() == cliente.strip().lower() == (sucursal_texto or "").strip().lower():
            continue

        estado_texto = (_texto(fila[idx["estado"]]) or "").strip().lower() if idx["estado"] is not None else ""
        estado = ESTADOS_MAP.get(estado_texto, "entregado")

        prefijo = _prefijo_folio(folio)
        sucursal_id = _encontrar_sucursal(prefijo, sucursal_texto, sucursales)

        fila_datos = {
            "folio": folio,
            "sucursal_texto_original": sucursal_texto,
            "sucursal_id": sucursal_id,
            "cliente_nombre": cliente,
            "cliente_telefono": _texto(fila[idx["telefono"]]) if idx["telefono"] is not None else None,
            "equipo": _texto(fila[idx["equipo"]]) if idx["equipo"] is not None else None,
            "marca": _texto(fila[idx["marca"]]) if idx["marca"] is not None else None,
            "modelo": _texto(fila[idx["modelo"]]) if idx["modelo"] is not None else None,
            "numero_serie": _texto(fila[idx["ns"]]) if idx["ns"] is not None else None,
            "garantia": _booleano_si_no(fila[idx["garantia"]]) if idx["garantia"] is not None else False,
            "falla_reportada": _texto(fila[idx["problema"]]) if idx["problema"] is not None else None,
            "fecha_recepcion": _fecha_iso(fila[idx["fecha_recepcion"]]) if idx["fecha_recepcion"] is not None else None,
            "autorizacion_precio": _booleano_si_no(fila[idx["autorizacion"]]) if idx["autorizacion"] is not None else None,
            "diagnostico": _texto(fila[idx["diagnostico"]]) if idx["diagnostico"] is not None else None,
            "folio_solicitud_traspaso": _texto(fila[idx["folio_traspaso"]]) if idx["folio_traspaso"] is not None else None,
            "fecha_entrega": _fecha_iso(fila[idx["fecha_salida"]]) if idx["fecha_salida"] is not None else None,
            "estado": estado,
            "items_costo": [],
        }

        costo_refaccion = fila[idx["costo_refaccion"]] if idx["costo_refaccion"] is not None else None
        servicio = fila[idx["servicio"]] if idx["servicio"] is not None else None
        paqueteria = fila[idx["paqueteria"]] if idx["paqueteria"] is not None else 0
        if costo_refaccion:
            fila_datos["items_costo"].append({"articulo": "Refacción", "costo": float(costo_refaccion)})
        if servicio:
            fila_datos["items_costo"].append({"articulo": "Servicio", "costo": float(servicio)})
        fila_datos["costo_paqueteria"] = float(paqueteria) if paqueteria else 0

        if folio in folios_vistos_en_este_archivo:
            fingerprint = _fingerprint(fila_datos)
            if fingerprint in folios_vistos_en_este_archivo[folio]:
                fila_datos["estatus"] = "duplicada"
            else:
                # Mismo número de folio, pero es claramente un trabajo
                # distinto (otra falla/diagnóstico/costo) — se conserva
                # con un folio único en vez de perderlo.
                sufijo = len(folios_vistos_en_este_archivo[folio]) + 1
                fila_datos["folio"] = f"{folio}-{sufijo}"
                fila_datos["estatus"] = "sin_sucursal" if sucursal_id is None else "lista"
                folios_vistos_en_este_archivo[folio].append(fingerprint)
        elif folio in folios_existentes:
            fila_datos["estatus"] = "duplicada"
            folios_vistos_en_este_archivo[folio] = [_fingerprint(fila_datos)]
        else:
            fila_datos["estatus"] = "sin_sucursal" if sucursal_id is None else "lista"
            folios_vistos_en_este_archivo[folio] = [_fingerprint(fila_datos)]

        filas_procesadas.append(fila_datos)

    return filas_procesadas
